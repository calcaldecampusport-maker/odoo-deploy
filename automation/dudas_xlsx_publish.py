#!/usr/bin/env python3
"""
Reads /tmp/dudas/<vat>_dudas.json (produced by dudas_xlsx_collect.py),
downloads the existing dudas_para_revisar.xlsx in each company root via SA,
merges new rows preserving user-filled `tu_decision`, uploads back.

Runs in /opt/automation/venv (has google + openpyxl).
"""
# === pipeline isolation guard (auto-injected) ===
import os as _os, sys as _sys
_HERE = _os.path.dirname(_os.path.abspath(__file__))
if _HERE not in _sys.path:
    _sys.path.insert(0, _HERE)
try:
    import companies as _comp_guard
    if getattr(_comp_guard, "PIPELINE_NAME", None) != 'cararjfam':
        raise RuntimeError(
            f"PIPELINE_MISMATCH: script {__file__} expected pipeline='cararjfam' "
            f"but loaded companies.PIPELINE_NAME={getattr(_comp_guard, 'PIPELINE_NAME', None)!r}"
        )
except ImportError:
    pass  # script sin dependencia de companies.py (e.g. drive_ops)
# === end isolation guard ===

import io
import json
import logging
import sys
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

sys.path.insert(0, _HERE)
import drive_ops  # noqa: E402

INPUT_DIR = Path("/tmp/dudas")
XLSX_NAME = "dudas_para_revisar.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HEADER = [
    "empresa", "tipo", "id_odoo", "ref_o_concepto", "fecha",
    "importe", "descripcion_corta", "motivo_duda", "sugerencia_actual",
    "tu_decision", "notas", "estado_actual", "primer_visto", "ultimo_visto",
]

log = logging.getLogger("dudas_publish")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def _row_key(r: dict) -> str:
    return f"{r.get('tipo','')}::{r.get('id_odoo','')}::{r.get('ref_o_concepto','')[:40]}"


def _row_to_list(r: dict, today: str, primer_visto: str = "") -> list:
    return [
        r.get("empresa", ""), r.get("tipo", ""), r.get("id_odoo", ""),
        r.get("ref_o_concepto", ""), r.get("fecha", ""), r.get("importe", ""),
        r.get("descripcion_corta", ""), r.get("motivo_duda", ""),
        r.get("sugerencia_actual", ""), r.get("tu_decision", ""),
        r.get("notas", ""), r.get("estado_actual", ""),
        primer_visto or today, today,
    ]


def _list_to_dict(values) -> dict:
    return {HEADER[i]: values[i] if i < len(values) else "" for i in range(len(HEADER))}


def _build_workbook(rows: list[list], periodic_data: dict | None = None, extractor_data: dict | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dudas"
    ws.append(HEADER)
    hdr_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    widths = {"A":25, "B":14, "C":9, "D":25, "E":12, "F":12, "G":30, "H":35, "I":35, "J":25, "K":40, "L":16, "M":12, "N":12}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"
    review_fill = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type="solid")
    descuadre_fill = PatternFill(start_color="fce4d6", end_color="fce4d6", fill_type="solid")
    for r in rows:
        ws.append(r)
        last = ws.max_row
        if (r[1] if len(r) > 1 else "") == "banco_descuadre":
            for c in ws[last]: c.fill = descuadre_fill
        elif (r[9] if len(r) > 9 else ""):
            for c in ws[last]: c.fill = review_fill
    if extractor_data:
        dups = extractor_data.get("duplicates") or []
        errs = extractor_data.get("errors") or []
        if dups:
            ws_d = wb.create_sheet("Duplicados")
            ws_d.append(["Archivo PDF", "Proveedor", "Ref factura", "Fecha", "Total", "Factura existente (id)"])
            for c in ws_d[1]:
                c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")
            dup_fill = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
            for it in dups:
                ws_d.append([it.get("file",""), it.get("supplier",""), it.get("ref",""),
                             str(it.get("invoice_date","")), it.get("total",""), it.get("invoice_id","")])
                for c in ws_d[ws_d.max_row]: c.fill = dup_fill
            for col, w in {"A":40,"B":30,"C":25,"D":12,"E":12,"F":15}.items():
                ws_d.column_dimensions[col].width = w
            ws_d.freeze_panes = "A2"
        if errs:
            # Preserve previous tu_decision/notas from existing Rechazados sheet
            prev_decisions = extractor_data.get("_prev_rechazos") or {}
            ws_e = wb.create_sheet("Rechazados")
            ws_e.append(["Archivo PDF", "Drive file_id", "Motivo del rechazo", "tu_decision", "notas"])
            for c in ws_e[1]:
                c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal="center")
            err_fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
            for it in errs:
                reason = (it.get("reason","") or "")[:1000]
                fid_key = it.get("id","")
                prev = prev_decisions.get(fid_key) or {}
                ws_e.append([it.get("file",""), fid_key, reason,
                             prev.get("tu_decision",""), prev.get("notas","")])
                for c in ws_e[ws_e.max_row]: c.fill = err_fill
            for col, w in {"A":40,"B":50,"C":80,"D":35,"E":30}.items():
                ws_e.column_dimensions[col].width = w
            ws_e.freeze_panes = "A2"

    if periodic_data and periodic_data.get('patterns'):
        ws2 = wb.create_sheet('Gastos_periodicos')
        ws2.append(['Tipo', 'Descripción', 'Cuenta', 'Signo', 'Frecuencia', 'Cada N días', 'Ocurrencias', 'Última fecha', 'Último importe', 'Importe medio', 'Esperada próxima', 'Días retraso', 'Falta'])
        for c in ws2[1]:
            c.fill = hdr_fill; c.font = hdr_font; c.alignment = Alignment(horizontal='center')
        miss_fill = PatternFill(start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
        ok_fill = PatternFill(start_color='e2efda', end_color='e2efda', fill_type='solid')
        for p in periodic_data['patterns']:
            ws2.append([
                p.get('tipo'), p.get('label'), p.get('account_code'), p.get('signo'),
                p.get('frecuencia'), p.get('avg_gap_dias'), p.get('ocurrencias'),
                p.get('ultima_fecha'), p.get('ultimo_importe'), p.get('importe_medio'),
                p.get('esperada_proxima'), p.get('dias_de_retraso'),
                'SI' if p.get('falta') else 'no',
            ])
            last = ws2.max_row
            for c in ws2[last]:
                c.fill = miss_fill if p.get('falta') else ok_fill
        for col, w in {'A':12,'B':35,'C':10,'D':10,'E':12,'F':12,'G':12,'H':14,'I':14,'J':14,'K':16,'L':12,'M':8}.items():
            ws2.column_dimensions[col].width = w
        ws2.freeze_panes = 'A2'
        # info row
        summ = periodic_data.get('summary') or {}
        info_row = [f"Análisis: {periodic_data.get('today','')}", f"Faltantes: {summ.get('faltantes',0)}", f"Detectados: {summ.get('total_periodicos_detectados',0)}"]
        ws2.append([])
        ws2.append(info_row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def find_xlsx(svc, folder_id: str):
    q = f"'{folder_id}' in parents and trashed=false and name='{XLSX_NAME}'"
    files = svc.files().list(q=q, fields="files(id,name)", supportsAllDrives=True).execute().get("files", [])
    return files[0] if files else None


def publish(svc, payload: dict) -> dict:
    folder = payload.get("pending_folder")
    if not folder:
        return {"company": payload["company"], "skipped": "no pending_folder"}
    file_meta = find_xlsx(svc, folder)
    if not file_meta:
        return {
            "company": payload["company"],
            "skipped": (
                f"file '{XLSX_NAME}' not found in {payload['company']} root. "
                "Crealo una vez (vacio) y el bot lo poblará."
            ),
        }

    existing_by_key = {}
    try:
        raw = svc.files().get_media(fileId=file_meta["id"], supportsAllDrives=True).execute()
        wb_existing = load_workbook(io.BytesIO(raw))
        ws_existing = wb_existing.active
        for row_cells in ws_existing.iter_rows(min_row=2, values_only=True):
            row_dict = _list_to_dict(list(row_cells))
            existing_by_key[_row_key(row_dict)] = row_dict
    except Exception as e:
        log.warning(f"could not read existing xlsx for {payload['company']}: {e}")

    today = payload["today"]
    fresh = payload["rows"]

    final_rows = []
    seen = set()
    new_count = 0
    preserved = 0
    for fr in fresh:
        key = _row_key(fr)
        seen.add(key)
        existing = existing_by_key.get(key)
        if existing:
            merged = dict(fr)
            decision = (existing.get("tu_decision") or "").strip()
            merged["tu_decision"] = decision
            primer = existing.get("primer_visto") or today
            user_notas = existing.get("notas") or ""
            if user_notas and len(user_notas) > len(fr.get("notas", "")):
                merged["notas"] = user_notas
            if decision: preserved += 1
            final_rows.append(_row_to_list(merged, today, str(primer)))
        else:
            final_rows.append(_row_to_list(fr, today))
            new_count += 1

    closed = 0
    for k, ex in existing_by_key.items():
        if k in seen: continue
        if (ex.get("tu_decision") or "").strip():
            closed += 1

    final_rows.sort(key=lambda r: (r[1] or "", str(r[2] or "")))
    # Load periodic patterns if recent JSON exists
    periodic_data = None
    try:
        from pathlib import Path as _Path
        import json as _json
        per_file = _Path(f"/tmp/periodic/{payload['vat']}_periodic.json")
        if per_file.exists():
            periodic_data = _json.loads(per_file.read_text())
    except Exception:
        log.exception('could not load periodic JSON')
    # Read existing xlsx Rechazados sheet to preserve tu_decision
    prev_rechazos = {}
    try:
        from openpyxl import load_workbook as _lw
        if existing_by_key or True:
            try:
                _raw = svc.files().get_media(fileId=file_meta["id"], supportsAllDrives=True).execute()
                _wb_old = _lw(io.BytesIO(_raw))
                if "Rechazados" in _wb_old.sheetnames:
                    _ws_old = _wb_old["Rechazados"]
                    for _r in _ws_old.iter_rows(min_row=2, values_only=True):
                        if _r and _r[1]:
                            prev_rechazos[_r[1]] = {
                                "tu_decision": _r[3] or "",
                                "notas": _r[4] or "",
                            }
            except Exception:
                pass
    except Exception:
        pass

    extractor_data = None
    try:
        from datetime import date as _date, timedelta as _td
        dups_acc, errs_acc = [], []
        # Look back 3 days
        for offset in range(0, 4):
            d = _date.today() - _td(days=offset)
            run_file = _Path(f"/tmp/extractor_runs/{d.isoformat()}.json")
            if not run_file.exists():
                continue
            run = _json.loads(run_file.read_text())
            for st in (run.get("summary") or []):
                if st.get("company") == payload.get("company"):
                    for dup in (st.get("duplicates") or []):
                        dup["_date"] = d.isoformat()
                        dups_acc.append(dup)
                    for err in (st.get("errors") or []):
                        err["_date"] = d.isoformat()
                        errs_acc.append(err)
        if dups_acc or errs_acc:
            extractor_data = {"duplicates": dups_acc, "errors": errs_acc, "_prev_rechazos": prev_rechazos}
    except Exception:
        log.exception("could not load extractor JSON")
    xlsx_bytes = _build_workbook(final_rows, periodic_data=periodic_data, extractor_data=extractor_data)

    from googleapiclient.http import MediaIoBaseUpload
    media = MediaIoBaseUpload(io.BytesIO(xlsx_bytes), mimetype=XLSX_MIME, resumable=False)
    svc.files().update(fileId=file_meta["id"], media_body=media, supportsAllDrives=True).execute()

    return {
        "company": payload["company"],
        "file_id": file_meta["id"],
        "rows_written": len(final_rows),
        "new": new_count,
        "preserved_decisions": preserved,
        "resolved": closed,
    }


def main():
    if not INPUT_DIR.exists():
        log.warning(f"no input dir at {INPUT_DIR}")
        return

    svc = drive_ops._service()
    out = []
    for f in sorted(INPUT_DIR.glob("*_dudas.json")):
        try:
            payload = json.loads(f.read_text())
            stats = publish(svc, payload)
        except Exception as e:
            log.exception(f"failed for {f.name}")
            stats = {"file": f.name, "error": str(e)}
        out.append(stats)
        log.info(f"{f.name}: {stats}")

    print(json.dumps(out, ensure_ascii=False))


if __name__ == "__main__":
    main()
