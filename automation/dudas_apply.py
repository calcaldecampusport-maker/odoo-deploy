#!/usr/bin/env python3
"""
Apply user decisions written in tu_decision column of dudas_para_revisar.xlsx.

This script:
  1. Downloads each company's xlsx (via SA — split: drive part)
  2. For each row with tu_decision filled, interprets common patterns and
     executes the matching action via Odoo ORM (split: odoo part)
  3. Marks estado_actual to one of:
       APERTURA, NETEADO, COMISION_BANCARIA, HIPOTECA, GASTO_NO_DEDUCIBLE,
       COBRO_FACTURA, RECONCILED_OPEN_LINE, PENDIENTE_HUMANO, ERROR
  4. Re-uploads xlsx

Run: python3 dudas_apply.py [--company-vat B...] [--dry-run]
"""
import argparse
import io
import json
import logging
import re
import sys
import subprocess
from pathlib import Path

sys.path.insert(0, "/opt/automation")
import drive_ops  # noqa: E402
import companies as comp  # noqa: E402

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

XLSX_NAME = "dudas_para_revisar.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HEADER = [
    "empresa", "tipo", "id_odoo", "ref_o_concepto", "fecha", "importe",
    "descripcion_corta", "motivo_duda", "sugerencia_actual",
    "tu_decision", "notas", "estado_actual", "primer_visto", "ultimo_visto",
]
ODOO_PYTHON = "/opt/odoo17/venv/bin/python"
APPLY_HELPER = "/opt/automation/dudas_apply_odoo.py"

log = logging.getLogger("dudas_apply")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def classify_decision(decision: str) -> dict:
    """Map a free-text decision to a structured action."""
    d = (decision or "").lower().strip()
    if not d:
        return {"action": "skip", "label": "vacio"}

    if any(k in d for k in ["apertura", "asiento de apertura", "asiento apertura"]):
        return {"action": "skip", "label": "APERTURA"}
    if any(k in d for k in ["ignorar", "saltar"]):
        return {"action": "skip", "label": "IGNORADO"}

    if any(k in d for k in ["neteado", "neteo", "se compensa", "se anula con", "compensado", "siguiente", "anterior"]):
        return {"action": "skip", "label": "NETEADO"}

    if any(k in d for k in ["comision banc", "comisión banc", "comision bancaria", "comisión bancaria"]):
        return {"action": "direct_entry", "account_code": "626000", "label": "COMISION_BANCARIA"}

    if any(k in d for k in ["hipoteca", "préstamo", "prestamo"]):
        return {"action": "direct_entry", "account_code": "520000", "label": "HIPOTECA"}

    if any(k in d for k in ["no deducible", "no-deducible", "non deducible", "no deduc"]):
        return {"action": "direct_entry", "account_code": "629000", "label": "GASTO_NO_DEDUCIBLE",
                "narration": "Gasto NO deducible: " + decision[:200]}

    if any(k in d for k in ["pago factura de ingres", "pago factura del cliente", "cobro factura"]):
        return {"action": "match_open_aml", "account_code": "430000", "label": "COBRO_FACTURA"}

    if "pago seguro" in d or "seguro anual" in d:
        return {"action": "direct_entry", "account_code": "625000", "label": "SEGURO"}

    if any(k in d for k in [
        "falta fact", "falta fra", "falta f.r.a", "pendiente fact", "queda como duda",
        "queda el moviemiento", "queda el movimiento", "mientras se sube", "espera factura",
    ]):
        return {"action": "skip", "label": "PENDIENTE_FACTURA"}

    if "pago nomina" in d or "pago nómina" in d:
        return {"action": "partial_reconcile_465", "label": "PARTIAL_RECONCILE"}

    if "saldo pendiente" in d or "queda pendiente" in d or "diferencia se queda" in d:
        return {"action": "partial_reconcile", "label": "PARTIAL_RECONCILE"}

    # User confirms the suggested AML in sugerencia_actual ("ok", "si", "confirmo")
    if d in ("ok", "si", "sí", "confirmo", "vale", "correcto") or d.startswith("ok "):
        return {"action": "confirm_proposal", "label": "PARTIAL_RECONCILE"}

    # Pago seguridad social (TGSS regimen general)
    if any(k in d for k in ["pago seguridad social", "pago ss", "seguridad social", "cotizacion ss", "tgss cotizacion"]):
        return {"action": "direct_entry", "account_code": "476000", "label": "PAGO_SS",
                "narration": "Pago Seg Social: " + decision[:200]}

    # Pago liquidacion IVA
    if any(k in d for k in ["pago iva", "liquidacion iva", "liquidación iva", "liquidacion de iva", "liquidación de iva", "autoliquidacion iva", "iva autoliquidacion", "iva mod 303", "modelo 303"]):
        return {"action": "direct_entry", "account_code": "477000", "label": "PAGO_IVA",
                "narration": "Pago liquidacion IVA: " + decision[:200]}

    # Pago retenciones IRPF (mod 111/115)
    if any(k in d for k in ["retenciones irpf", "pago irpf", "retenciones e ing", "retenciones a cta", "mod 111", "mod 115", "retenciones e ingresos a cuenta"]):
        return {"action": "direct_entry", "account_code": "475100", "label": "PAGO_IRPF",
                "narration": "Pago retenciones IRPF: " + decision[:200]}

    # Pago alquiler (arrendamiento)
    if any(k in d for k in ["pago alquiler", "alquiler mensual", "renta alquiler", "arrendamiento"]):
        return {"action": "match_open_aml_or_direct", "account_code": "621000", "label": "PAGO_ALQUILER",
                "narration": "Pago alquiler: " + decision[:200]}

    # Pago factura proveedor — try to match against open AML on 410
    if any(k in d for k in ["pago factura", "pago fra", "pago de fra", "pago de factura",
                            "agrupacion de facturas", "agrupación de facturas",
                            "agrupacion facturas", "factura rectificativa", "factgura rectificativa"]):
        return {"action": "match_open_aml", "account_code": "410000", "label": "PAGO_FACTURA",
                "narration": "Pago factura proveedor: " + decision[:200]}

    # Cobros TPV / liquidaciones — cliente paga via TPV
    if "liquidacion efectuada" in d or "liquidaci" in d and "efectuada" in d or "cobro tpv" in d:
        return {"action": "direct_entry", "account_code": "430000", "label": "COBRO_TPV",
                "narration": "Cobro TPV: " + decision[:200]}

    # Gympass
    if "gympass" in d:
        return {"action": "direct_entry", "account_code": "430000", "label": "COBRO_GYMPASS",
                "partner_name": "Gympass US LLC",
                "narration": "Cobro Gympass: " + decision[:200]}

    # Comisiones bancarias / SEPA / TPV / devoluciones bancarias
    if any(k in d for k in [
        "gastos devolucion", "gastos devoluciones",
        "comisiones banc", "comision banc",
        "comsiones banc",
        "tarifa plana",
        "emision sepa", "emisi" "sepa",
        "emision remesa sepa",
        "liquidacion por emision",
        "liquidaci" "n por emisi",
    ]):
        return {"action": "direct_entry", "account_code": "626000", "label": "COMISION_BANCARIA",
                "narration": decision[:200]}

    # Devolucion de recibos cliente
    if "devolucion de recibo" in d or "devoluci" in d and "recibo" in d:
        return {"action": "direct_entry", "account_code": "430000", "label": "DEVOLUCION_CLIENTE",
                "narration": "Devolucion recibo SEPA: " + decision[:200]}

    # Emision Remesa SEPA: cobro masivo de clientes
    if "emision remesa" in d or "emisi" in d and "remesa" in d or "abona la cuenta de clientes" in d:
        return {"action": "direct_entry", "account_code": "430000", "label": "COBRO_REMESA_SEPA",
                "narration": "Cobro remesa SEPA: " + decision[:200]}

    return {"action": "human", "label": "PENDIENTE_HUMANO"}


def collect_rows(svc, cfg) -> tuple[str, list[dict], list]:
    folder = cfg.get("pending_folder")
    q = f"'{folder}' in parents and trashed=false and name='{XLSX_NAME}'"
    files = svc.files().list(q=q, fields="files(id,name)", supportsAllDrives=True).execute().get("files", [])
    if not files:
        return None, [], []
    fid = files[0]["id"]
    raw = svc.files().get_media(fileId=fid, supportsAllDrives=True).execute()
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows_with_decision = []
    for idx, row in enumerate(all_rows):
        if row is None or row[0] is None:
            continue
        d = (row[9] or "").strip() if row[9] else ""
        if not d:
            continue
        rows_with_decision.append({
            "row_index": idx,
            "empresa": row[0], "tipo": row[1], "id_odoo": row[2],
            "ref_o_concepto": row[3], "fecha": row[4], "importe": row[5],
            "descripcion_corta": row[6], "motivo_duda": row[7],
            "sugerencia_actual": row[8], "tu_decision": d,
            "notas": row[10] or "", "estado_actual": row[11] or "",
            "primer_visto": row[12] or "", "ultimo_visto": row[13] or "",
        })
    return fid, rows_with_decision, all_rows


def write_xlsx(rows_data: list[list]) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Dudas"
    ws.append(HEADER)
    hdr_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for col, w in {"A":25,"B":14,"C":9,"D":25,"E":12,"F":12,"G":30,"H":35,"I":35,"J":25,"K":40,"L":18,"M":12,"N":12}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    applied_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
    pending_fill = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")
    error_fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
    for r in rows_data:
        ws.append(r)
        last = ws.max_row
        estado = (r[11] if len(r) > 11 else "") or ""
        if estado.startswith("ERROR"):
            for c in ws[last]: c.fill = error_fill
        elif estado in ("APERTURA","IGNORADO","NETEADO","COMISION_BANCARIA","HIPOTECA","GASTO_NO_DEDUCIBLE","COBRO_FACTURA","RECONCILED_OPEN_LINE","PARTIAL_RECONCILE","SEGURO","PENDIENTE_FACTURA"):
            for c in ws[last]: c.fill = applied_fill
        elif estado == "PENDIENTE_HUMANO":
            for c in ws[last]: c.fill = pending_fill
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def upload_xlsx(svc, fid: str, content: bytes):
    from googleapiclient.http import MediaIoBaseUpload
    media = MediaIoBaseUpload(io.BytesIO(content), mimetype=XLSX_MIME, resumable=False)
    svc.files().update(fileId=fid, media_body=media, supportsAllDrives=True).execute()


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--company-vat")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    svc = drive_ops._service()
    overall = []

    for cfg in comp.COMPANIES:
        if args.company_vat and cfg["vat"] != args.company_vat:
            continue
        log.info(f"=== {cfg['name']} ===")
        fid, decisions, all_rows = collect_rows(svc, cfg)
        if not fid:
            log.warning(f"no xlsx for {cfg['name']}")
            continue

        # Classify decisions
        actions_payload = []
        for row in decisions:
            cls = classify_decision(row["tu_decision"])
            actions_payload.append({
                "row_index": row["row_index"],
                "tipo": row["tipo"], "id_odoo": row["id_odoo"],
                "importe": row["importe"], "decision": row["tu_decision"],
                "action": cls["action"], "label": cls["label"],
                "account_code": cls.get("account_code"),
                "narration": cls.get("narration"),
                "partner_name": cls.get("partner_name"),
            })

        log.info(f"  {len(decisions)} decisions -> classified")
        for a in actions_payload[:20]:
            log.info(f"    row#{a['row_index']} tipo={a['tipo']} id={a['id_odoo']} -> {a['label']}")

        # Hand off to Odoo helper to execute
        if not args.dry_run:
            tmp_in = Path(f"/tmp/dudas/{cfg['vat']}_actions.json")
            tmp_out = Path(f"/tmp/dudas/{cfg['vat']}_actions_result.json")
            tmp_in.parent.mkdir(parents=True, exist_ok=True)
            tmp_in.write_text(json.dumps({
                "company_id": cfg["odoo_company_id"],
                "company_vat": cfg["vat"],
                "actions": actions_payload,
            }, ensure_ascii=False))
            try:
                result = subprocess.run(
                    [ODOO_PYTHON, APPLY_HELPER, "--input", str(tmp_in), "--output", str(tmp_out)],
                    capture_output=True, text=True, timeout=600,
                )
                log.info(f"  helper rc={result.returncode}")
                if result.stderr:
                    log.info(f"  stderr tail: {result.stderr[-400:]}")
            except subprocess.TimeoutExpired:
                log.error("helper timed out")
            try:
                exec_results = json.loads(tmp_out.read_text())
            except Exception:
                exec_results = []
        else:
            exec_results = []

        results_by_idx = {r["row_index"]: r for r in exec_results}

        # Update xlsx rows in-place
        new_rows_data = []
        for idx, original in enumerate(all_rows):
            if original is None or original[0] is None:
                continue
            row_list = list(original)
            while len(row_list) < len(HEADER):
                row_list.append("")
            executed = results_by_idx.get(idx)
            if executed:
                row_list[11] = executed.get("estado_actual") or row_list[11]
                if executed.get("note"):
                    existing = row_list[10] or ""
                    row_list[10] = (existing + " | " + executed["note"])[:500] if existing else executed["note"][:500]
            new_rows_data.append(row_list)

        if not args.dry_run:
            upload_xlsx(svc, fid, write_xlsx(new_rows_data))

        overall.append({
            "company": cfg["name"],
            "decisions": len(decisions),
            "executed": len(exec_results),
            "labels": {a["label"]: sum(1 for x in actions_payload if x["label"] == a["label"]) for a in actions_payload},
        })

    print(json.dumps(overall, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
