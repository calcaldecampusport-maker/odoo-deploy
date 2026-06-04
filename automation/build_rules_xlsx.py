#!/usr/bin/env python3
"""
Genera REGLAS_SISTEMA.xlsx con todas las reglas de negocio + validaciones del
sistema CARARJFAM. Pensado para que otro agente Claude (o un humano) lo lea y
entienda todo el comportamiento sin tener que abrir el código.

Estructura: 1 hoja por área del sistema + README inicial.
Cada fila = una regla con: ID, Regla, Lógica/Fórmula, Severidad, Fuente, Script, Notas.
"""
# === pipeline isolation guard (auto-injected) ===
import os as _os, sys as _sys
_HERE = _os.path.dirname(_os.path.abspath(__file__))
if _HERE not in _sys.path:
    _sys.path.insert(0, _HERE)
try:
    import companies as _comp_guard
    if getattr(_comp_guard, "PIPELINE_NAME", None) != 'cararjfam':
        raise RuntimeError(
            f"PIPELINE_MISMATCH: script {__file__} expected pipeline='cararjfam' "
            f"but loaded companies.PIPELINE_NAME={getattr(_comp_guard, 'PIPELINE_NAME', None)!r}"
        )
except ImportError:
    pass  # script sin dependencia de companies.py (e.g. drive_ops)
# === end isolation guard ===

import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

sys.path.insert(0, _HERE)
from drive_ops import _service
from googleapiclient.http import MediaFileUpload

DRIVE_PARENT = "1RZjKO1GqJuPURl6WTsl2R9egwm7cyYFQ"  # carpeta "Mi Odoo CARARJFAM"
DRIVE_FILE_ID = "1lX032XqZ63u73jM6wDlIZsekqeQ-pzkG"  # REGLAS_SISTEMA.xlsx (pre-creado vía OAuth user; SA solo UPDATE)
OUT_PATH = "/tmp/REGLAS_SISTEMA.xlsx"

HDR_FILL = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
ERROR_FILL = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
WARN_FILL = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type="solid")
INFO_FILL = PatternFill(start_color="d9e1f2", end_color="d9e1f2", fill_type="solid")
THIN = Side(border_style="thin", color="999999")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

SEVERITY_COLORS = {"error": ERROR_FILL, "warning": WARN_FILL, "info": INFO_FILL, "critical": ERROR_FILL}

COLS = ["ID", "Regla", "Lógica / Fórmula", "Severidad", "Empresa", "Fuente legal/Spec", "Script donde aplica", "Notas"]
COL_WIDTHS = {"A": 8, "B": 35, "C": 60, "D": 12, "E": 18, "F": 30, "G": 35, "H": 50}


def _new_sheet(wb, title: str, intro: str = ""):
    ws = wb.create_sheet(title)
    if intro:
        ws.append([intro])
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLS))
        c = ws.cell(row=1, column=1)
        c.font = Font(italic=True, color="555555")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[1].height = 35
    ws.append(COLS)
    for cell in ws[ws.max_row]:
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    for col, w in COL_WIDTHS.items():
        ws.column_dimensions[col].width = w
    return ws


def _add_rule(ws, rule):
    ws.append([
        rule["id"], rule["regla"], rule["formula"], rule["severidad"].upper(),
        rule.get("empresa", "TODAS"),
        rule["fuente"], rule["script"], rule.get("notas", ""),
    ])
    fill = SEVERITY_COLORS.get(rule["severidad"].lower(), INFO_FILL)
    for cell in ws[ws.max_row]:
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BORDER
    ws.row_dimensions[ws.max_row].height = max(30, 15 * (1 + rule["formula"].count("\n") + rule["regla"].count("\n")))


def build():
    wb = Workbook()
    # README
    ws0 = wb.active
    ws0.title = "README"
    ws0.column_dimensions["A"].width = 22
    ws0.column_dimensions["B"].width = 120

    readme = [
        ("Documento", "REGLAS_SISTEMA.xlsx — biblioteca completa de reglas del pipeline contable CARARJFAM"),
        ("Versión", "2026-05-18"),
        ("Cobertura", "Nóminas (validaciones SS), Facturas, Bancos, Dudas, Aprendizaje, Extractor, Backup"),
        ("Para quién", "Lo leen agentes Claude que retomen el proyecto, o humanos que auditen el sistema"),
        ("Formato", "Una hoja por área. Cada fila = una regla con ID único, lógica/fórmula, severidad, fuente legal o spec, script donde se aplica, notas."),
        ("Severidades", "CRITICAL = aborta operación. ERROR = aborta operación. WARNING = no aborta, anota aviso. INFO = informativo, sin acción."),
        ("Color filas", "Rojo = error/critical. Amarillo = warning. Azul = info."),
        ("Convención IDs", "<ÁREA><N>: NOM = nóminas, FAC = facturas, BNK = bancos, DUD = dudas, APR = aprendizaje, EXT = extractor, BAK = backup, PER = gastos periódicos."),
        ("Columna Empresa", "Indica qué empresa creó/usa la regla. Valores: GLOBAL = regla universal Spanish law/PGC/infra (no se debería divergir entre empresas) · TODAS = aplica a las 3 empresas pero configurable · CARARJFAM, BT, AUSTRAL = específica de esa empresa · CARARJFAM,BT (separadas por coma) = aplica a varias específicas. Si dos empresas necesitan reglas contradictorias, duplicar la fila y poner cada empresa en su versión."),
        ("Fuente actualizada", "Archivo se actualiza cada vez que el agente añade/cambia una regla; última edición arriba."),
        ("Servidor de referencia", "212.227.40.122 (IONOS). Odoo 17 Community. BD cararjfam (prod) y cararjfam_test (sandbox)."),
        ("Documentos relacionados", "RECOVERY.md (runbook), BLUEPRINT.md (patrones reutilizables), HANDOFF.md / HANDOFF_AUSTRAL.md (traspasos)."),
        ("Sources legales clave", "LGSS RD-leg 8/2015, Reglamento Cotización RD 2064/1995, Orden ISM/118/2026 (tipos 2026), PGC PYMES España, AEB43 norma CSB43."),
    ]
    for k, v in readme:
        ws0.append([k, v])
        ws0.cell(row=ws0.max_row, column=1).font = Font(bold=True)
        ws0.cell(row=ws0.max_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    # ====== Para Claude (onboarding completo del sistema) ======
    ws_c = wb.create_sheet("Para_Claude")
    ws_c.column_dimensions["A"].width = 28
    ws_c.column_dimensions["B"].width = 130
    titulo = ws_c.cell(row=1, column=1, value="Onboarding para otro agente Claude - leeme primero")
    titulo.font = Font(bold=True, size=14, color="1f4e79")
    ws_c.merge_cells("A1:B1")
    ws_c.row_dimensions[1].height = 28

    onboarding = [
        ("Que hace este sistema",
         "Pipeline contable nocturno autonomo: cada noche el cron del VPS recoge documentos (facturas, nominas, IRPF, SS, extractos bancarios) de carpetas Drive de las empresas, los clasifica con un LLM (Claude headless), los procesa y crea asientos en Odoo 17 Community via ORM. Bancos se concilian en 3 pasadas (1:1, subset-sum 1:N, near-match). Lo que el sistema no resuelve solo lo escribe en un xlsx dudas_para_revisar.xlsx en Drive; el usuario rellena la columna tu_decision con texto libre, el bot lee, clasifica y ejecuta via ORM. Cada decision humana se convierte en learned.rule auto-aplicable a lineas futuras. Email diario con resumen + xlsx adjunto."),

        ("Empresas en produccion",
         "BD cararjfam: empresa 1 = CARARJFAM2019,SL (CIF B93653392, tiene socios autonomos: Carlos Alcalde, Concepcion Arjona, con salario_especie). Empresa 2 = BEST TRAINING RINCON DE LA VICTORIA SL (CIF B72349137, trabajadores normales). BD cararjfam_test: sandbox + empresa 3 = AUSTRAL (CIF ESB44821965, usuario contabilidad@austral.es). BD demo: vacia, sin uso."),

        ("Servidor",
         "IONOS Ubuntu 24.04, IP 212.227.40.122, dominios erp.carajfam.com (prod) y demo.carajfam.com (test/austral). SSH como root con clave ed25519 en ~/.ssh/odoo_carajfam (PasswordAuthentication off, fail2ban activo). Odoo 17 en /opt/odoo17/odoo, config /etc/odoo17.conf, data_dir /opt/odoo17/.local/share/Odoo. Custom addon learned_rules en /opt/odoo17/custom-addons/. Master Odoo BD: ver admin_passwd en odoo17.conf."),

        ("Patron critico - dos venvs",
         "NUNCA mezclar libs en el mismo venv: /opt/odoo17/venv (Odoo + psycopg2 + pyOpenSSL 21) vs /opt/automation/venv (Google API + pandas + openpyxl + xlrd + csb43). Instalar google-api-python-client en el venv de Odoo rompe pyOpenSSL.X509StoreFlags y tumba Odoo. Los scripts se comunican entre venvs por JSON en /tmp + subprocess (dudas_apply.py llama a dudas_apply_odoo.py)."),

        ("Pipeline diaria (cron user odoo, hora Madrid)",
         "23:23 learning_drive (descarga reglas CSV de Drive) -> 23:25 learning (importa reglas + scan pasivo facturas posted) -> 23:30 extractor (LLM clasifica PDFs y enruta) -> 23:35 poller (legacy, idle) -> 23:36 dudas_apply (lee xlsx tu_decision, ejecuta via ORM) -> 23:37 apply_rules_to_bank (aplica TODAS las learned.rule a lineas no conciliadas) + bank_reconciler --threshold 90 -> 23:38 bank_multi_reconciler (subset-sum) + dudas_xlsx_collect (recoge dudas -> JSON) -> 23:39 dudas_xlsx_publish (sube xlsx a Drive) -> 23:40 email_summary (HTML + xlsx adjunto). Extra: 04:00 backup_to_drive (rolling 7 dias); 08:00 los dias 10 de ene/abr/jul/oct: periodic_expenses_check."),

        ("Drive - Service Account + workaround quota=0",
         "SA email: vps-odoo-automation@carajfam-automation.iam.gserviceaccount.com (creds en /etc/automation_sa.json). Tiene quota 0 en Drive personal del usuario; NO puede CREATE archivos nuevos, solo UPDATE existentes. Workaround: archivos como dudas xlsx, RECOVERY.md, REGLAS_SISTEMA.xlsx y los 7 backup_DIA.tar.gz se PRE-CREARON via OAuth del usuario humano (file_ids fijos hardcodeados en codigo). La SA solo los actualiza. Si necesitas crear un archivo nuevo, usa el MCP create_file de Drive con cuenta del usuario, no la SA. Carpetas Drive por empresa: CARARJFAM 'Mi Odoo CARARJFAM' (1RZjKO1GqJuPURl6WTsl2R9egwm7cyYFQ), BT (en companies.py:pending_folder), AUSTRAL 'Mi Odoo AUSTRAL' (1SNDTko-SgeYNjyJ-_635ObprBDVWm-Jd). Subcarpetas estandar por empresa: Pendientes (raiz, donde sube docs el usuario) + Contabilizado odoo + revision + Aprendizajes_aplicados + ignorados (creada solo cuando rechazo se marca como ignorar)."),

        ("Documentos hermanos - leelos en este orden si arrancas desde 0",
         "1) /opt/automation/RECOVERY.md (sync en Drive file_id 1cYLjkrXgkJxKS-2nBD8GX0DFugMLznZe) - runbook completo de reinstalacion + lista de datos de configuracion clave. Imprescindible. 2) BLUEPRINT.md (en repo local odoo-deploy/) - patrones reutilizables si vas a montar algo similar en otro dominio. 3) HANDOFF.md / HANDOFF_AUSTRAL.md - como portar el sistema a otro ERP / anadir una empresa nueva. 4) ARQUITECTURA_CARAJFAM.md - explicacion didactica del sistema para humanos no-tecnicos. 5) Memoria del proyecto Claude: ~/.claude/projects/.../memory/MEMORY.md y archivos referenciados (preferencias usuario + reglas operativas, como 'pregunta una a una' y 'actualiza RECOVERY.md en cada cambio estructural')."),

        ("Como regenero ESTE xlsx (REGLAS_SISTEMA.xlsx)",
         "sudo -u odoo /opt/automation/venv/bin/python /opt/automation/build_rules_xlsx.py. El script lee las reglas hardcodeadas en su propia code, construye xlsx con openpyxl en /tmp, sube via SA UPDATE al file_id fijo 1lX032XqZ63u73jM6wDlIZsekqeQ-pzkG. Si anades reglas nuevas, edita el script (seccion rules_<area>.append({...})) y vuelve a ejecutar."),

        ("Reglas operativas no negociables (memoria del proyecto)",
         "1) Pregunta UNA cosa a la vez al usuario (Carlos), no listas de 5 puntos numerados; no le gustan. 2) Cada cambio ESTRUCTURAL del VPS (nuevo script, nuevo cron, nuevo addon, nueva BD, nuevo subdomain, nuevo modulo, cambio de path) -> actualizar RECOVERY.md en local + scp al VPS + UPDATE en Drive (file_id 1cYLjkrXgkJxKS-2nBD8GX0DFugMLznZe) en el MISMO turno. Si no, en 6 meses el RECOVERY esta obsoleto y el restore no funciona cuando se necesita. 3) Si una regla aprendida (learned.rule) tiene caso edge especifico de una empresa, marcarla con company_id correcto; NO mezclar reglas entre empresas (los socios autonomos de CARARJFAM no aplican a BT). 4) Idempotencia: 'already reconciled' / 'already exists' es exito, NO error. Pipelines re-pasan sin doblar acciones."),

        ("Pitfalls universales del proyecto",
         "1) Total negativo factura proveedor = factura rectificativa (in_refund), no in_invoice negativo (Odoo rechaza). 2) Chart l10n_es trae 12 IVAs al mismo tipo (G/S/EX/EU/IG/RC/ND); filtrar prefijos especiales o el calculo IVA falla. 3) En CARARJFAM las cuentas 475100 / 476000 son 6 digitos, no 4 (4751/4760 NO existen). 4) Cuentas 465/475/476/477 requieren reconcile=True; tras restore hay que habilitarlo manualmente. 5) Bank reconciler: nunca confiar SOLO en importe; filtrar por similitud partner/concepto, sino cruza Culligan 72,88 con MediaMarkt 30E (caso real). 6) Lineas contactless TPV: solo match 100% exacto contra factura; NUNCA partial ni near-match (regla usuario tras falsos positivos). 7) Cursor cerrado tras cr.commit() fuera del 'with' -> InterfaceError; capturar valores ORM en variables locales antes. 8) HEIC images (iPhone): Claude headless no las lee bien; convertir a JPG primero. 9) Gmail SMTP requiere app-password 16 chars (Security -> App Passwords), NO la pwd normal. 10) Charts l10n_es PYMES en Odoo 17 cambio name (campo translatable) y account_type; tras restore validar que account_type sea correcto (asset_cash, liability_current, etc)."),

        ("Como entrar a debuggear sin romper nada",
         "Para QUERY/READ: sudo -u postgres psql -d cararjfam -c 'SELECT ...' o via ORM con sudo -u odoo /opt/odoo17/venv/bin/python -c '...' importando odoo.api.Environment. Para PROBAR cambios sin afectar prod: usa BD cararjfam_test (clon del estado actual + cron interno desactivado + SMTP off; no envia emails reales). Para reset rapido cararjfam_test: pg_dump cararjfam | pg_restore cararjfam_test (script en RECOVERY.md seccion 16)."),

        ("Como se notifica al usuario",
         "Email diario a c.alcalde.campusport@gmail.com via Gmail SMTP (app password 16 chars en /opt/automation/email_config.py). Contiene: lista facturas creadas hoy (link a Odoo), seccion Documentos rechazados como duplicado, seccion de conciliaciones bancarias propuestas/near-matches, xlsx adjunto por empresa. El usuario rellena la columna tu_decision del xlsx y al dia siguiente el cron procesa esa decision."),

        ("Conceptos contables espanoles esenciales",
         "PGC PYMES Espana: 400/410 proveedores, 430 clientes, 465 remuneraciones pdtes pago empleados, 475100 HP retenciones IRPF (6 digitos!), 476000 Org SS acreedores, 520 prestamos corto plazo, 572 banco, 600 compras, 625 seguros, 626 servicios bancarios, 629 otros (placeholder no deducible), 640 sueldos, 642 SS empresa, 477 HP IVA repercutido. Asiento de nomina (spec usuario): DR 640=bruto total devengo + DR 642=aport empresa SS / CR 4751=IRPF + CR 476=aport_empresa+ss_trab+especie_socio + CR 465.NNN per trabajador=liquido_cash. Tipos cotizacion SS 2026: CC empresa 24,35% / desempleo empresa indef 5,50% temp 6,70% / FP empresa 0,60% / FOGASA 0,20% / CC trabajador 4,85% / desempleo trabajador 1,55% / FP trabajador 0,10%. BCCC empleado DEBE = BCCC empresa (LGSS 147)."),

        ("Para entender ESTE xlsx",
         "9 hojas. README = metadatos. Para_Claude = esta hoja (onboarding). Resto = una hoja por area del sistema (Nominas, Facturas, Bancos, Dudas, Aprendizaje, Extractor, Backup, Gastos_periodicos). Cada fila = una regla con ID unico (NOM01, FAC02, etc), severidad coloreada (rojo=error/critical, amarillo=warning, azul=info), columna Empresa (GLOBAL/TODAS/CARARJFAM/BT/AUSTRAL/combinaciones). Si dos empresas necesitan reglas opuestas, duplicar fila con cada Empresa correspondiente."),
    ]
    row = 2
    section_fill = PatternFill(start_color="e7f0fa", end_color="e7f0fa", fill_type="solid")
    for titulo_sec, contenido in onboarding:
        c1 = ws_c.cell(row=row, column=1, value=titulo_sec)
        c1.font = Font(bold=True, color="1f4e79", size=11)
        c1.fill = section_fill
        c1.alignment = Alignment(vertical="top", wrap_text=True)
        c2 = ws_c.cell(row=row, column=2, value=contenido)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        c2.font = Font(size=10)
        ws_c.row_dimensions[row].height = max(40, min(330, 15 + len(contenido) // 130 * 14))
        row += 1
    ws_c.freeze_panes = "A2"

    # ====== Nóminas ======
    ws = _new_sheet(wb, "Nominas",
        "Asiento contable de nómina y 10 validaciones SS empresa española. Aplica a "
        "nomina_processor.py. El asiento siempre debe cuadrar (Σ DR == Σ CR) o se aborta.")
    rules_nom = [
        # Estructura del asiento
        {"empresa": "GLOBAL", "id": "NOM01", "regla": "Asiento nómina — estructura cuentas PGC",
         "formula": "DR 640 = Σ bruto (total devengo)\nDR 642 = Σ ss_empresa (aport total: CC + AT + FOGASA + FP + desempleo empresa)\nCR 4751 = Σ irpf retenido\nCR 476 = Σ ss_empresa + Σ ss_trabajador + Σ especie_socio\nCR 465.NNN per trabajador = bruto_i − ss_trab_i − irpf_i − especie_i",
         "severidad": "info", "fuente": "PGC PYMES España + spec usuario",
         "script": "nomina_processor.py:process()",
         "notas": "El especie de socios autónomos CARARJFAM va al CR 476 (cuota autónomo a TGSS), NO al DR 642 (eso descuadraría)."},
        {"empresa": "GLOBAL", "id": "NOM02", "regla": "Subcuenta 465 por trabajador",
         "formula": "Para cada empleado: buscar 465* cuyo name contenga su NIF. Si no existe, crear 465NNN (NNN incremental empezando 001) con nombre 'Liquido pdte pago <empleado> (<NIF>)'. Reusable.",
         "severidad": "info", "fuente": "spec usuario (mantener trazabilidad pago por trabajador)",
         "script": "nomina_processor.py:_get_or_create_465_for_employee",
         "notas": "Permite conciliar pagos bancarios contra el 465 específico del empleado."},
        {"empresa": "GLOBAL", "id": "NOM03", "regla": "Math nómina per empleado",
         "formula": "liquido_i == bruto_i − ss_trab_i − irpf_i − especie_i  (± 0,05€)",
         "severidad": "error", "fuente": "aritmética PGC",
         "script": "nomina_processor.py:process()",
         "notas": "Si no cuadra → rc=10 → mueve PDF a /revision/ sin postear."},
        {"empresa": "GLOBAL", "id": "NOM04", "regla": "Asiento global cuadra",
         "formula": "Σ DR == Σ CR  (± 0,05€)",
         "severidad": "error", "fuente": "principio partida doble",
         "script": "nomina_processor.py:process() — chequeo pre-post",
         "notas": "Si no cuadra → rc=10 → no postea."},
        {"empresa": "GLOBAL", "id": "NOM05", "regla": "Idempotencia por ref",
         "formula": "ref = f'Nomina {YYYY-MM} ({company_id})'. Si ya existe → rc=20 (duplicate, mueve a contabilizado, no recrea).",
         "severidad": "info", "fuente": "diseño pipeline",
         "script": "nomina_processor.py:process()",
         "notas": "Evita doble posteo si se vuelve a subir el mismo PDF."},
        # 10 validaciones SS
        {"empresa": "GLOBAL", "id": "NOM-R1", "regla": "BCCC empleado == BCCC empresa",
         "formula": "bccc_empleado == bccc_empresa  (± 0,05€)",
         "severidad": "error", "fuente": "LGSS art. 147 — base de cotización es UNA",
         "script": "nomina_processor.py:_run_nomina_ss_validations",
         "notas": "Si difieren → ERROR EN NÓMINA: gestor metió base distinta para empresa (caso real FRANCO POLO marzo 2026, sobrepago ~95€). Se anota en narración."},
        {"empresa": "GLOBAL", "id": "NOM-R2", "regla": "BCCC ≈ devengo (±2%)",
         "formula": "|bccc − total_devengo| ≤ devengo × 0,02",
         "severidad": "warning", "fuente": "art. 23 Reglamento Cotización RD 2064/1995",
         "script": "nomina_processor.py:_run_nomina_ss_validations",
         "notas": "El 2% permite dietas/km exentos. Si supera, revisar exentos o detectar error de cálculo."},
        {"empresa": "GLOBAL", "id": "NOM-R3", "regla": "Base AT/EP ≈ BCCC",
         "formula": "|base_at − bccc| ≤ 0,05€",
         "severidad": "warning", "fuente": "art. 24.1 Reglamento Cotización",
         "script": "nomina_processor.py:_run_nomina_ss_validations",
         "notas": "Diferentes solo si hay horas extras estructurales (suben la base AT)."},
        {"empresa": "GLOBAL", "id": "NOM-R4", "regla": "Cuota CC empresa correcta",
         "formula": "cuota_cc_empresa == round(bccc_empresa × 0,2435, 2)",
         "severidad": "error", "fuente": "Orden ISM/118/2026 (tipos 2026)",
         "script": "nomina_processor.py:_run_nomina_ss_validations",
         "notas": "Tipo CC empresa régimen general 24,35%."},
        {"empresa": "GLOBAL", "id": "NOM-R5", "regla": "Cuota desempleo empresa correcta",
         "formula": "cuota_des_empresa == bccc × 5,50% (indef)  o  × 6,70% (temp)",
         "severidad": "warning", "fuente": "Orden ISM/118/2026",
         "script": "nomina_processor.py:_run_nomina_ss_validations",
         "notas": "Discriminar por tipo_contrato extraído del PDF."},
        {"empresa": "GLOBAL", "id": "NOM-R6", "regla": "Cuota FP empresa correcta",
         "formula": "cuota_fp == bccc × 0,60%",
         "severidad": "error", "fuente": "Orden ISM/118/2026",
         "script": "nomina_processor.py:_run_nomina_ss_validations", "notas": ""},
        {"empresa": "GLOBAL", "id": "NOM-R7", "regla": "Cuota FOGASA correcta",
         "formula": "cuota_fogasa == bccc × 0,20%",
         "severidad": "error", "fuente": "Orden ISM/118/2026",
         "script": "nomina_processor.py:_run_nomina_ss_validations", "notas": ""},
        {"empresa": "GLOBAL", "id": "NOM-R8", "regla": "Topes BCCC por grupo cotización",
         "formula": "bccc ∈ [min_grupo, max_grupo]",
         "severidad": "info", "fuente": "Anexo Orden anual TGSS",
         "script": "(pendiente implementar — necesita tabla anual)",
         "notas": "Para 2026 min grupo 1 ~1.490€, max ~4.909,50€. Cambia cada año."},
        {"empresa": "GLOBAL", "id": "NOM-R9", "regla": "Jornada parcial coherente",
         "formula": "bccc_parcial ≈ bccc_full × (horas_contratadas / 40)",
         "severidad": "info", "fuente": "RD 2317/2002",
         "script": "(pendiente implementar — necesita datos contrato)",
         "notas": "Detectaría error tipo: gestor usa base jornada completa para parcial."},
        {"empresa": "GLOBAL", "id": "NOM-R10", "regla": "Σ cuotas calculadas ≈ aport_empresa_total",
         "formula": "|Σ(cc + at + des + fp + fogasa) − aport_empresa_total| ≤ 0,10€",
         "severidad": "warning", "fuente": "aritmética",
         "script": "nomina_processor.py:_run_nomina_ss_validations",
         "notas": "Cuadre global del aport empresa contra suma de cuotas extraídas."},
    ]
    for r in rules_nom: _add_rule(ws, r)

    # ====== Facturas ======
    ws = _new_sheet(wb, "Facturas",
        "Reglas para extracción y creación de facturas (in_invoice / in_refund / out_invoice). "
        "Aplica a process_invoice.py + extractor.py.")
    rules_fac = [
        {"id": "FAC01", "regla": "Confianza umbral auto-post",
         "formula": "if extraction_confidence >= 0.9 → action_post(); else state='draft' → entra dudas xlsx",
         "severidad": "info", "fuente": "configuración usuario",
         "script": "process_invoice.py", "notas": ""},
        {"empresa": "GLOBAL", "id": "FAC02", "regla": "Total negativo → factura rectificativa",
         "formula": "if total < 0 → move_type='in_refund', líneas en valor absoluto",
         "severidad": "critical", "fuente": "Odoo rechaza in_invoice con total negativo",
         "script": "process_invoice.py", "notas": "Sin esta detección Odoo no permite postear."},
        {"empresa": "GLOBAL", "id": "FAC03", "regla": "Detección duplicados",
         "formula": "exists(partner_id, ref, invoice_date, company_id) → rc=20, mover a contabilizado sin crear",
         "severidad": "info", "fuente": "diseño pipeline",
         "script": "process_invoice.py:already_exists",
         "notas": "Evita doble posteo si usuario re-sube mismo PDF."},
        {"empresa": "GLOBAL", "id": "FAC04", "regla": "Math factura cuadra",
         "formula": "|subtotal + tax_total − total| ≤ 0,05€",
         "severidad": "error", "fuente": "aritmética",
         "script": "extractor.py:_validate",
         "notas": "Si no cuadra → revisión manual."},
        {"empresa": "GLOBAL", "id": "FAC05", "regla": "Filtro impuestos especiales l10n_es",
         "formula": "SPECIAL_TAX_PREFIXES = (' EX', ' EU', ' IG', ' RC', ' ND'). Excluir al elegir IVA default.",
         "severidad": "critical", "fuente": "chart l10n_es trae 12 IVAs al mismo tipo",
         "script": "process_invoice.py:_select_tax",
         "notas": "Sin filtro coge 'EX' (extranjero) y neutraliza IVA por error."},
        {"empresa": "CARARJFAM", "id": "FAC06", "regla": "VAT correction (aprendizaje)",
         "formula": "Antes de validar VAT: consultar learned.rule(rule_type='vat_correction', pattern in supplier_name). Si encuentra → reemplazar VAT extraído por notes de la regla.",
         "severidad": "info", "fuente": "aprendizaje desde xlsx Rechazados",
         "script": "process_invoice.py:_maybe_correct_vat",
         "notas": "Permite saltar VAT validation cuando gestor escribe CIF correcto en la hoja Rechazados."},
        {"empresa": "GLOBAL", "id": "FAC07", "regla": "Cuenta default por doc_type",
         "formula": "invoice→600000, nomina→640000, irpf_payment→475100, ss_payment→642000, other_official→629000",
         "severidad": "info", "fuente": "PGC PYMES",
         "script": "process_invoice.py:DOC_TYPE_DEFAULT_ACCOUNT", "notas": ""},
        {"empresa": "GLOBAL", "id": "FAC08", "regla": "Búsqueda partner por VARIANTES de VAT",
         "formula": "Antes de crear partner, search por: VAT raw, VAT canonical (con ES), VAT sin ES. Si cualquiera matchea → reutilizar partner existente (no crear duplicado).",
         "severidad": "critical", "fuente": "post-mortem GANESHA (B86002318 vs ESB86002318 → 2 partners)",
         "script": "process_invoice.py:find_or_create_supplier",
         "notas": "normalize_vat puede producir formas distintas según el input; buscar solo por la canonical pierde matches."},
        {"empresa": "GLOBAL", "id": "FAC09", "regla": "Fallback por nombre normalizado",
         "formula": "Si no encuentra por VAT (variantes), buscar partner por nombre normalizado (upper + collapse spaces + strip SL/SA suffix). Si matchea → reutilizar y actualizar VAT.",
         "severidad": "critical", "fuente": "post-mortem MATEO MOTOR (mismo VAT pero 2 partners por race)",
         "script": "process_invoice.py:find_or_create_supplier + _norm_partner_name",
         "notas": "Detecta duplicados causados por race / typo VAT / VAT ausente. Solo crea nuevo partner si nada matchea."},
        {"empresa": "GLOBAL", "id": "FAC10", "regla": "NUNCA SQL crudo para campos validados (vat, email)",
         "formula": "SIEMPRE usar partner.write({'vat': X}). Raw SQL UPDATE res_partner SET vat=... salta normalize_vat de base_vat y produce duplicados.",
         "severidad": "critical", "fuente": "post-mortem GANESHA (yo causé el duplicado con SQL crudo)",
         "script": "dudas_apply_odoo._create_vat_correction (corregido)",
         "notas": "Si base_vat rechaza checksum → with_context(no_vat_validation=True).write() como fallback."},
        {"empresa": "GLOBAL", "id": "FAC11", "regla": "Cron diario detector duplicados",
         "formula": "Cron 06:00 diario ejecuta detect_duplicate_partners.py: 3 SQL detectan VAT-idéntico, nombre-normalizado-idéntico, VAT-igual-salvo-prefijo-ES. Si encuentra → email alerta.",
         "severidad": "warning", "fuente": "safety net post-mortem",
         "script": "detect_duplicate_partners.py + cron 0 6 * * *",
         "notas": "Detecta cualquier duplicado que se cuele a pesar de FAC08+FAC09 (race, manual UI input, restore)."},
        {"empresa": "GLOBAL", "id": "FAC12", "regla": "Comprobar SIEMPRE si ya está contabilizado ANTES de contabilizar (incl. importaciones y cross-diario)",
         "formula": "Antes de crear cualquier asiento buscar duplicado: (a) por nº/ref del documento en TODA la company, no solo el diario destino; (b) cross-diario: el nº puede existir ya en OV (migración Sage, ref 'SAGE-xxxx / Doc <nº>') → buscar 'Doc <nº>'; (c) extractos por (company, journal, periodo, nº líneas); (d) sin nº fiable, por (partner + importe + fecha). Contar y reportar lo omitido; nunca contabilizar a ciegas.",
         "severidad": "critical", "fuente": "incidente mayo 2026: 327 ventas duplicadas vs asientos Sage (diario OV)",
         "script": "process_invoice.py / importadores masivos / bank_importer.py",
         "notas": "Generaliza FAC03 (que solo cubre re-subida del mismo PDF). Comprobar-antes-de-contabilizar es PARTE de contabilizar, en todas las empresas y también en importaciones masivas."},
        {"empresa": "AUSTRAL", "id": "FAC13", "regla": "Retención IRPF en facturas (alquileres / profesionales)",
         "formula": "Si hay retención: total = base + IVA − IRPF. Detectar irpf_rate/irpf_amount; aplicar impuesto de compra reutilizable amount=−rate%. Cuenta: 19% (alquiler) → 475100003; 15%/7% (profesional) → 475100002.",
         "severidad": "error", "fuente": "facturas de alquiler/servicios con retención; PGC 4751",
         "script": "process_invoice.py:find_or_create_irpf_tax + validate_payload; extractor PROMPT irpf_rate/irpf_amount",
         "notas": "Sin esto el cuadre base+IVA=total falla y la factura va a Revisión. Origen: alquiler ANTONIO BELLIDO."},
        {"empresa": "AUSTRAL", "id": "FAC14", "regla": "IVA incluido en tickets de caja → reescalar a base neta",
         "formula": "Si las líneas suman ~total (con IVA) y NO el subtotal, reescalar cada línea a neto (por tax_rate de la línea; fallback ratio subtotal/Σlíneas) y cuadrar el último céntimo.",
         "severidad": "error", "fuente": "tickets caja euros con IVA incluido (líneas brutas)",
         "script": "process_invoice.py:_normalize_iva_included_lines",
         "notas": "Conservador: solo actúa si Σlíneas cuadra con el total, no con el subtotal."},
        {"empresa": "AUSTRAL", "id": "FAC15", "regla": "Ventas factura-a-factura (libro de facturas emitidas)",
         "formula": "Cliente por NIF (vat ∈ [nif, ESnif]; fallback por nombre) → su cuenta 430+código. Diario INV(28). Preservar nº oficial como move.name. IVA por importe (21/10/4) y 0% por país: UE→intracom(506), resto/Canarias→export(507). Abonos = out_refund con importes en positivo.",
         "severidad": "error", "fuente": "export ERP comercial 'Facturas/Abonos emitidas'",
         "script": "importador masivo de ventas",
         "notas": "ANTES aplicar FAC12: muchas ya están en asientos Sage (diario OV, ref 'Doc <nº>'). No duplicar."},
    ]
    for r in rules_fac: _add_rule(ws, r)

    # ====== Bancos ======
    ws = _new_sheet(wb, "Bancos",
        "3 pasadas de matching/conciliación bancaria + reglas contactless + propagación.")
    rules_bnk = [
        {"empresa": "GLOBAL", "id": "BNK01", "regla": "Pasada 1 — matching 1:1",
         "formula": "Para cada bank.statement.line, buscar AML cuyo score ≥ 90. Score = amount(50) + direction(10) + partner_name(25) + ref(15) + date_proximity(5).",
         "severidad": "info", "fuente": "diseño pipeline",
         "script": "bank_reconciler.py + bank_matcher._score", "notas": ""},
        {"empresa": "GLOBAL", "id": "BNK02", "regla": "Pasada 2 — subset-sum 1:N",
         "formula": "Si no encontró match 1:1: buscar subconjunto de AMLs (size 2..12) cuya suma cuadre con bank line ± 0,10€. Usado para IRPF trimestral (Σ mensuales).",
         "severidad": "info", "fuente": "diseño pipeline",
         "script": "bank_multi_reconciler.py:_find_subset", "notas": ""},
        {"empresa": "GLOBAL", "id": "BNK03", "regla": "Pasada 3 — near-match (no auto-acción)",
         "formula": "Diferencia 0,5€ < diff ≤ 100€ + similitud partner/concepto ≥ 0,3 → surfacear en email para revisión humana.",
         "severidad": "info", "fuente": "diseño pipeline",
         "script": "bank_matcher.find_near_matches_for_company", "notas": ""},
        {"empresa": "CARARJFAM,BT", "id": "BNK04", "regla": "Contactless solo match 100% exacto",
         "formula": "Si payment_ref startswith 'TRANSACCION CONTACTLESS': solo buscar facturas con amount_total dentro de ±0,005€. Nunca near-match, nunca subset, nunca learned.rule.",
         "severidad": "critical", "fuente": "regla usuario (evita falsos positivos por importe coincidente)",
         "script": "bank_matcher.py:_is_contactless + bank_multi_reconciler",
         "notas": "Caso real: bsl Culligan 72,88€ se cruzaba mal con MediaMarkt 30€."},
        {"empresa": "GLOBAL", "id": "BNK05", "regla": "Pre-filtro near-match: partner similarity",
         "formula": "Excluir near-match si: name_similarity < 0,3 AND first_word(partner) NOT in concept AND ref NOT in concept.",
         "severidad": "warning", "fuente": "post-mortem falsos positivos",
         "script": "bank_matcher.find_near_matches_for_company", "notas": ""},
        {"empresa": "GLOBAL", "id": "BNK06", "regla": "Apply learned rules — all-words matching",
         "formula": "pattern.split() ⊆ payment_ref.upper().split()  (todas las palabras del pattern presentes en concepto, no necesariamente contiguas)",
         "severidad": "info", "fuente": "post-mortem 'LIQUIDACION EMISION' vs 'Liquidacion Por Emision'",
         "script": "apply_rules_to_bank.py",
         "notas": "Substring estricto fallaba con conectores. All-words tolerante."},
        {"empresa": "GLOBAL", "id": "BNK07", "regla": "Idempotencia reconciliación",
         "formula": "'already reconciled' se trata como éxito, NO error. Re-pasadas no doblan acciones.",
         "severidad": "info", "fuente": "principio idempotency pipelines",
         "script": "todos los reconcilers", "notas": ""},
        {"empresa": "AUSTRAL", "id": "BNK08", "regla": "Lectura robusta de extractos .xls/.xlsx",
         "formula": "xls: xlrd DIRECTO (bypass del check de versión pandas↔xlrd), fallback HTML y openpyxl. Cabecera detectada hasta fila 25 (BBVA: F.VALOR/F.CONTABLE). Fechas con hora se recortan a YYYY-MM-DD. Nº de cuenta CCC desnudo (16-24 díg sin prefijo ES) como IBAN hint. NO_ES_BANCO: rechazar exports de facturas (FACTURA/CLIENTE/NIF sin SALDO).",
         "severidad": "error", "fuente": "formatos reales ABANCA/BBVA/Cajamar",
         "script": "bank_importer.py:_read_xls_robusto/_find_iban_in_text/_parse_date_es",
         "notas": "Implementado solo en el pipeline austral."},
        {"empresa": "AUSTRAL", "id": "BNK10", "regla": "Todo extracto importado se concilia (auto + pendientes con aprendizaje)",
         "formula": "Al importar un extracto, bank_importer llama cmd_auto_reconcile: aplica learned.rule(bank, conf>=0.85) a las lineas. Las que no casan quedan en la pestana web Revision>Conciliaciones pendientes con info del banco + propuesta (bank_matcher) + fiabilidad. Al resolver, si learn=1, se crea/actualiza learned.rule(bank) para auto-conciliar futuros movimientos parecidos.",
         "severidad": "error", "fuente": "regla usuario: siempre conciliar al importar",
         "script": "bank_importer.py:import_file + conciliacion_ops.py + web /api/conciliacion",
         "notas": "Bucle de aprendizaje: cada decision manual mejora la auto-conciliacion siguiente. Reglas gestionables a mano en la web (Conciliaciones pendientes > Ver reglas): ver/crear/editar/activar learned.rule(bank) via /api/conciliacion/reglas."},
        {"empresa": "AUSTRAL", "id": "BNK09", "regla": "Journal de banco filtrado por company",
         "formula": "_find_journal busca solo journals de la company del pipeline (_AUSTRAL_COMPANY_ID=4). Un IBAN duplicado en otra company (3 'AUSTRAL en migración') NO debe capturar el statement.",
         "severidad": "critical", "fuente": "incidente: Caja Rural se contabilizó en company 3 por IBAN duplicado",
         "script": "bank_importer.py:_find_journal",
         "notas": "Evita contaminar otra empresa con extractos de Austral."},
    ]
    for r in rules_bnk: _add_rule(ws, r)

    # ====== Dudas xlsx ======
    ws = _new_sheet(wb, "Dudas",
        "Patrones de classify_decision en dudas_apply.py. Texto libre del usuario → acción programática.")
    rules_dud = [
        {"id": "DUD01", "regla": "apertura / asiento de apertura", "formula": "→ skip APERTURA",
         "severidad": "info", "fuente": "uso usuario", "script": "dudas_apply.classify_decision", "notas": "Línea pre-fundacional."},
        {"id": "DUD02", "regla": "neteado / siguiente / anterior / compensado", "formula": "→ skip NETEADO",
         "severidad": "info", "fuente": "uso usuario", "script": "dudas_apply.classify_decision", "notas": "Par cancelado, no acción contable."},
        {"id": "DUD03", "regla": "comisión banc / tarifa plana / emisión SEPA", "formula": "→ DR 626 COMISION_BANCARIA",
         "severidad": "info", "fuente": "PGC PYMES",
         "script": "dudas_apply.classify_decision", "notas": ""},
        {"id": "DUD04", "regla": "hipoteca / préstamo", "formula": "→ DR 520 HIPOTECA",
         "severidad": "info", "fuente": "PGC PYMES", "script": "dudas_apply.classify_decision", "notas": ""},
        {"id": "DUD05", "regla": "no deducible", "formula": "→ DR 629 GASTO_NO_DEDUCIBLE + nota 'no deducible'",
         "severidad": "info", "fuente": "Ley IS", "script": "dudas_apply.classify_decision", "notas": ""},
        {"id": "DUD06", "regla": "pago factura / cobro factura / pago fra", "formula": "→ match_open_aml 410/430",
         "severidad": "info", "fuente": "uso usuario", "script": "dudas_apply.classify_decision", "notas": ""},
        {"id": "DUD07", "regla": "pago seguro / seguro anual", "formula": "→ DR 625 SEGURO",
         "severidad": "info", "fuente": "PGC", "script": "dudas_apply.classify_decision", "notas": ""},
        {"id": "DUD08", "regla": "falta fact / pendiente fact / mientras se sube", "formula": "→ skip PENDIENTE_FACTURA",
         "severidad": "info", "fuente": "uso usuario", "script": "dudas_apply.classify_decision", "notas": ""},
        {"id": "DUD09", "regla": "pago nomina", "formula": "→ partial_reconcile vs 465 abierto (líquido empleado)",
         "severidad": "info", "fuente": "uso usuario", "script": "dudas_apply.classify_decision", "notas": ""},
        {"id": "DUD10", "regla": "ok / si / confirmo", "formula": "→ confirm_proposal: parsear sugerencia 'apunte X = Y' y reconciliar vs ese AML (preferencia cuenta 465>410>430>475>476)",
         "severidad": "info", "fuente": "uso usuario",
         "script": "dudas_apply_odoo._create_vat_correction… (confirm_proposal handler)", "notas": ""},
        {"empresa": "CARARJFAM", "id": "DUD11", "regla": "pago Seguridad Social / TGSS cotizacion", "formula": "→ DR 476 PAGO_SS",
         "severidad": "info", "fuente": "PGC", "script": "dudas_apply.classify_decision", "notas": ""},
        {"empresa": "CARARJFAM", "id": "DUD12", "regla": "pago IVA / liquidacion IVA / mod 303", "formula": "→ DR 477 PAGO_IVA",
         "severidad": "info", "fuente": "PGC", "script": "dudas_apply.classify_decision", "notas": ""},
        {"empresa": "CARARJFAM", "id": "DUD13", "regla": "retenciones IRPF / pago IRPF / mod 111 / mod 115", "formula": "→ DR 475100 PAGO_IRPF",
         "severidad": "info", "fuente": "PGC", "script": "dudas_apply.classify_decision", "notas": ""},
        {"id": "DUD14", "regla": "pago alquiler / arrendamiento", "formula": "→ match_open_aml 410 PAGO_ALQUILER",
         "severidad": "info", "fuente": "uso usuario", "script": "dudas_apply.classify_decision", "notas": ""},
        {"id": "DUD15", "regla": "no hacer / buscaré la contrapartida / manualmente", "formula": "→ skip PENDIENTE_USUARIO",
         "severidad": "info", "fuente": "uso usuario", "script": "dudas_apply.classify_decision",
         "notas": "Usuario gestiona manualmente, no tocar."},
        {"id": "DUD16", "regla": "subida fra / subida factura / factura subida", "formula": "→ smart_subida_match: buscar in_invoice abierta ±0,05€ con partner first_word en concepto. 1 candidato → reconcilia. 0 o múltiples → queda pendiente.",
         "severidad": "info", "fuente": "uso usuario",
         "script": "dudas_apply.classify_decision + dudas_apply_odoo smart_subida_match",
         "notas": "Confía en usuario: solo verifica que factura abierta exista."},
        {"id": "DUD17", "regla": "Rechazos sheet — borrar/eliminar", "formula": "→ trash en Drive (move to trash)",
         "severidad": "warning", "fuente": "uso usuario",
         "script": "dudas_apply._do_drive_action",
         "notas": "Reversible desde papelera Drive 30 días."},
        {"id": "DUD18", "regla": "Rechazos sheet — ignorar/omitir/saltar", "formula": "→ mover a subcarpeta /ignorados/",
         "severidad": "info", "fuente": "uso usuario", "script": "dudas_apply._do_drive_action", "notas": ""},
        {"id": "DUD19", "regla": "Rechazos sheet — reprocesar/intentar de nuevo", "formula": "→ mover de /revision/ a /Pendientes/ (cron 23:30 reintenta)",
         "severidad": "info", "fuente": "uso usuario", "script": "dudas_apply._do_drive_action", "notas": ""},
        {"id": "DUD20", "regla": "Rechazos sheet — CIF: <vat> (aprendizaje)",
         "formula": "Detecta regex r'\\b[A-Z]?\\d{7,8}[A-Z]?\\b' + keyword 'cif|nif|vat|es el'. → crea learned.rule(rule_type=vat_correction, pattern=partner_name) + actualiza partner si existe + mueve PDF a Pendientes/. Próxima vez extractor sustituye VAT automáticamente.",
         "severidad": "info", "fuente": "aprendizaje activo del usuario",
         "script": "dudas_apply.classify_rechazo_decision + dudas_apply_odoo._create_vat_correction",
         "notas": "Sistema aprende correcciones VAT del usuario. Caso real GANESHA marzo 2026."},
    ]
    for r in rules_dud: _add_rule(ws, r)

    # ====== Aprendizaje ======
    ws = _new_sheet(wb, "Aprendizaje",
        "Modelo custom learned.rule (addon learned_rules). 3 tipos de regla, auto-aprende desde decisiones.")
    rules_apr = [
        {"empresa": "GLOBAL", "id": "APR01", "regla": "rule_type='bank'", "formula": "pattern (concepto) → account_id (cuenta destino) + opcional partner_id",
         "severidad": "info", "fuente": "diseño pipeline",
         "script": "learned_rules + apply_rules_to_bank.py",
         "notas": "Reglas aplicadas a líneas bancarias sin conciliar. Match 'all-words present'."},
        {"empresa": "GLOBAL", "id": "APR02", "regla": "rule_type='invoice'", "formula": "pattern (descripción línea) → account_id (cuenta gasto)",
         "severidad": "info", "fuente": "diseño pipeline",
         "script": "learned_rules + learning.py",
         "notas": "Sugiere cuenta al posteo de factura según descripción línea."},
        {"empresa": "GLOBAL", "id": "APR03", "regla": "rule_type='vat_correction'", "formula": "pattern (partner_name uppercase) → notes (VAT correcto)",
         "severidad": "info", "fuente": "creado fase 2 hoja Rechazados",
         "script": "learned_rules + process_invoice._maybe_correct_vat",
         "notas": "Consultado antes de validación VAT en process_invoice. Sustituye VAT extraído del PDF por el guardado."},
        {"empresa": "GLOBAL", "id": "APR04", "regla": "Propagación al aplicar bank rule",
         "formula": "Al ejecutar direct_entry en dudas: derivar pattern de 2 palabras del concepto (saltar stopwords). Crear learned.rule(source=passive). Aplicar a TODAS las líneas no conciliadas que matcheen el pattern.",
         "severidad": "info", "fuente": "diseño aprendizaje activo",
         "script": "dudas_apply_odoo._propagate_to_similar",
         "notas": "Una decisión usuario → N reconciliaciones automáticas."},
        {"empresa": "GLOBAL", "id": "APR05", "regla": "Source rules: active / passive / system",
         "formula": "active = creado manual CSV. passive = auto-aprendido desde xlsx. system = regla intrínseca (TGSS AUTONOMOS).",
         "severidad": "info", "fuente": "diseño addon", "script": "learned_rules/models", "notas": ""},
        {"empresa": "GLOBAL", "id": "APR06", "regla": "Confianza umbral apply_rules_to_bank",
         "formula": "Sólo aplica reglas con confidence >= 0.85",
         "severidad": "info", "fuente": "diseño pipeline", "script": "apply_rules_to_bank.py:RULE_CONFIDENCE_MIN", "notas": ""},
    ]
    for r in rules_apr: _add_rule(ws, r)

    # ====== Extractor ======
    ws = _new_sheet(wb, "Extractor",
        "Clasificación document_type via Claude headless + routing al processor adecuado.")
    rules_ext = [
        {"empresa": "GLOBAL", "id": "EXT01", "regla": "document_type values",
         "formula": "Uno de: invoice, nomina, irpf_payment, ss_payment, other_official, not_a_document",
         "severidad": "info", "fuente": "PROMPT extractor",
         "script": "extractor.py:PROMPT", "notas": ""},
        {"empresa": "GLOBAL", "id": "EXT02", "regla": "Routing por doc_type",
         "formula": "invoice → process_invoice.py\nnomina → nomina_processor.py\nirpf_payment / ss_payment / other_official → tax_payment_processor.py\nbank_statement → bank_importer.py\nsepa → ignore",
         "severidad": "info", "fuente": "diseño pipeline",
         "script": "extractor.py", "notas": ""},
        {"empresa": "GLOBAL", "id": "EXT03", "regla": "Auto-post threshold by type",
         "formula": "invoice: confidence >= 0,9 auto-post (sino draft)\nnomina / irpf_payment / ss_payment / other_official: SIEMPRE post (documentos autoritativos, math verificable)",
         "severidad": "info", "fuente": "configuración usuario",
         "script": "process_invoice.py + nomina_processor.py", "notas": ""},
        {"empresa": "GLOBAL", "id": "EXT04", "regla": "Validación math factura",
         "formula": "|subtotal + tax_total − total| ≤ 0,05€",
         "severidad": "error", "fuente": "aritmética",
         "script": "extractor.py:_validate",
         "notas": "Si no cuadra → revision/."},
        {"empresa": "GLOBAL", "id": "EXT05", "regla": "Validación math nómina (employee)",
         "formula": "bruto − irpf − ss − especie = liquido (per empleado, ± 0,05€)",
         "severidad": "error", "fuente": "aritmética PGC",
         "script": "extractor.py:_validate + nomina_processor.py",
         "notas": ""},
        {"empresa": "GLOBAL", "id": "EXT06", "regla": "Skip filenames",
         "formula": "SKIP_FILENAME_HINTS = ('dudas','aprendizaje','_aplicado','_procesado')",
         "severidad": "info", "fuente": "diseño",
         "script": "extractor.py", "notas": "Evita reprocesar artefactos del pipeline."},
        {"empresa": "GLOBAL", "id": "EXT07", "regla": "Move file after process",
         "formula": "rc=0 → Contabilizado odoo/\nrc=20 (duplicate) → Contabilizado odoo/\notros rc → revision/",
         "severidad": "info", "fuente": "diseño pipeline",
         "script": "extractor.py", "notas": "PDF nunca se queda en Pendientes/ tras procesar."},
        {"empresa": "GLOBAL", "id": "EXT09", "regla": "NUNCA sys.path.insert hacia OTRA carpeta automation",
         "formula": "En cada script de /opt/automation/X.py el primer sys.path.insert DEBE apuntar a /opt/automation. Nunca a /opt/automation_austral o similar — eso provocaría que import companies cargue el companies.py incorrecto y procese empresas incorrectas.",
         "severidad": "critical", "fuente": "post-mortem bug 30-may a 1-jun 2026 (CARARJFAM/BT 3 dias sin procesar porque extractor.py leia companies.py de austral)",
         "script": "extractor.py + cualquier processor multi-tenant",
         "notas": "Si se copia codigo entre /opt/automation y /opt/automation_austral, revisar SIEMPRE los sys.path.insert y rutas absolutas. Bug-class: copy-paste cruzado entre pipelines hermanos."},
        {"empresa": "GLOBAL", "id": "EXT11", "regla": "Aislamiento entre pipelines multi-empresa (4 capas)",
         "formula": "1) Cada companies.py declara PIPELINE_NAME + DB_NAME + EXPECTED_VATS. 2) Cada script hace sys.path.insert(0, dirname(__file__)) → carga companies.py de SU misma carpeta. 3) Asserts comp.PIPELINE_NAME esperado al arrancar, aborta con RuntimeError(PIPELINE_MISMATCH) si carga el companies.py incorrecto. 4) Cron usa cd /opt/automation_X explícito antes de cada comando.",
         "severidad": "critical", "fuente": "refuerzo post-mortem EXT09 (bug sys.path cruzado entre pipelines)",
         "script": "companies.py (metadata) + cada *.py (preamble guard auto-inyectado) + crontab",
         "notas": "El guard usa __file__ para resolver la carpeta real del script, asi se mantiene robusto independientemente del cwd. Re-aplicar con _isolate_pipelines.py es idempotente (no duplica el bloque)."},
        {"empresa": "GLOBAL", "id": "EXT10", "regla": "Carpeta Pendientes limpia — solo PDFs/imagenes de facturas",
         "formula": "SKIP_FILENAME_HINTS contiene: 'dudas','aprendizaje','_aplicado','_procesado','backup_','recovery'. Cualquier archivo en Pendientes/ que NO sea factura debe ir a subcarpeta 'Sistema (no procesar)' o ser excluido por patron.",
         "severidad": "warning", "fuente": "post-mortem bug 1-jun 2026 (backups y RECOVERY.md acumulados en Pendientes de CARARJFAM)",
         "script": "extractor.py:SKIP_FILENAME_HINTS",
         "notas": "El backup_to_drive.py mantiene los 7 file_ids fijos pero independientes de la ubicacion; se pueden mover sin romper. Igual RECOVERY.md (file_id fijo)."},
        {"empresa": "GLOBAL", "id": "EXT08", "regla": "Persistencia resultado diario",
         "formula": "Cada noche escribe /tmp/extractor_runs/<YYYY-MM-DD>.json con summary (por empresa: duplicates[], errors[], created[]).",
         "severidad": "info", "fuente": "lectura por email + xlsx Rechazados",
         "script": "extractor.py final",
         "notas": "dudas_xlsx_publish lee últimos 4 días para hoja Rechazados/Duplicados."},
    ]
    for r in rules_ext: _add_rule(ws, r)

    # ====== Backup ======
    ws = _new_sheet(wb, "Backup",
        "Política de backup rolling 7 días al Drive vía Service Account.")
    rules_bak = [
        {"empresa": "GLOBAL", "id": "BAK01", "regla": "Rolling 7 días por día semana",
         "formula": "DAY_FILE_IDS = {0:LUNES, 1:MARTES, ..., 6:DOMINGO}. Cada noche sobrescribe el archivo del weekday actual. SA UPDATE (no CREATE — cuota personal Drive = 0).",
         "severidad": "info", "fuente": "limitación SA Drive personal",
         "script": "backup_to_drive.py", "notas": ""},
        {"empresa": "GLOBAL", "id": "BAK02", "regla": "Contenido del backup",
         "formula": "Un único tar.gz con: db.dump (pg_dump -Fc), filestore.tar.gz, custom-addons.tar.gz, automation.tar.gz (sin venv), configs.tar.gz (/etc/odoo17.conf + nginx + letsencrypt), secrets.tar.gz (SA JSON + email config), crontab_odoo.txt, RECOVERY.md.",
         "severidad": "info", "fuente": "diseño DR",
         "script": "backup_to_drive.py:build_full_backup", "notas": ""},
        {"empresa": "GLOBAL", "id": "BAK03", "regla": "Sudoers para pg_dump",
         "formula": "/etc/sudoers.d/odoo_pgdump → 'odoo ALL=(postgres) NOPASSWD: /usr/bin/pg_dump'",
         "severidad": "info", "fuente": "necesario para que user odoo dispare pg_dump como user postgres",
         "script": "Restore manual", "notas": ""},
        {"empresa": "GLOBAL", "id": "BAK04", "regla": "Horario cron",
         "formula": "0 4 * * * (04:00 diario, no choca con pipeline 23:23-23:40)",
         "severidad": "info", "fuente": "diseño",
         "script": "crontab user odoo", "notas": ""},
        {"empresa": "GLOBAL", "id": "BAK05", "regla": "RECOVERY.md sincronización triple",
         "formula": "Cada cambio estructural: editar local C:\\Users\\pc\\Documents\\odoo-deploy\\RECOVERY.md → scp a /opt/automation/RECOVERY.md → svc.files().update(fileId='1cYLjkrXgkJxKS-...').",
         "severidad": "critical", "fuente": "memoria proyecto project_carajfam_recovery_maintenance.md",
         "script": "manual + agente Claude",
         "notas": "Si se omite, RECOVERY.md no reflejará el sistema actual y restore fallará."},
    ]
    for r in rules_bak: _add_rule(ws, r)

    # ====== Análisis trimestral ======
    ws = _new_sheet(wb, "Gastos_periodicos",
        "Análisis trimestral (10 ene/abr/jul/oct) de patrones recurrentes y detección de faltantes.")
    rules_per = [
        {"empresa": "GLOBAL", "id": "PER01", "regla": "Cron disparador",
         "formula": "0 8 10 1,4,7,10 * (08:00 días 10 de ene/abr/jul/oct)",
         "severidad": "info", "fuente": "configuración usuario", "script": "periodic_expenses_check.py", "notas": ""},
        {"empresa": "GLOBAL", "id": "PER02", "regla": "Detección patrón periódico",
         "formula": "count >= 3 ocurrencias && CV gaps < 0,35 (coefficient of variation 35%) → es patrón regular",
         "severidad": "info", "fuente": "estadística", "script": "periodic_expenses_check.py:analyze_company", "notas": ""},
        {"empresa": "GLOBAL", "id": "PER03", "regla": "Marcado como faltante",
         "formula": "today > expected_next + avg_gap * 0,5 → MISSING (50% del gap como gracia)",
         "severidad": "warning", "fuente": "configuración usuario", "script": "periodic_expenses_check.py", "notas": ""},
        {"empresa": "GLOBAL", "id": "PER04", "regla": "Salida — hoja extra xlsx",
         "formula": "/tmp/periodic/<vat>_periodic.json → dudas_xlsx_publish añade hoja 'Gastos_periodicos' (rojo si falta, verde si OK)",
         "severidad": "info", "fuente": "diseño", "script": "dudas_xlsx_publish.py + periodic_expenses_check.py", "notas": ""},
        {"empresa": "GLOBAL", "id": "PER05", "regla": "Aprendizaje automático",
         "formula": "Sin lista predefinida — el sistema detecta patrones del histórico (LOOKBACK_MONTHS=18). Cuanto más histórico, más detecta.",
         "severidad": "info", "fuente": "diseño usuario", "script": "periodic_expenses_check.py", "notas": ""},
    ]
    for r in rules_per: _add_rule(ws, r)

    wb.save(OUT_PATH)
    return OUT_PATH


def upload_to_drive(path: str):
    """UPDATE el archivo pre-creado. La SA no puede CREATE (cuota=0); el archivo
    debe existir creado vía OAuth user (una sola vez)."""
    svc = _service()
    media = MediaFileUpload(path, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", resumable=False)
    res = svc.files().update(
        fileId=DRIVE_FILE_ID, media_body=media,
        fields="id,name,size,modifiedTime", supportsAllDrives=True,
    ).execute()
    print(f"updated REGLAS_SISTEMA.xlsx id={DRIVE_FILE_ID}: {res}")
    return DRIVE_FILE_ID


if __name__ == "__main__":
    path = build()
    print(f"built {path}")
    try:
        fid = upload_to_drive(path)
        print(f"drive file_id={fid}")
    except Exception as e:
        print(f"upload failed (will create via MCP fallback): {e}")
