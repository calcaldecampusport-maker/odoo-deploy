#!/usr/bin/env python3
"""
Apply user decisions written in tu_decision column of dudas_para_revisar.xlsx.

This script:
  1. Downloads each company's xlsx (via SA — split: drive part)
  2. For each row with tu_decision filled, interprets common patterns and
     executes the matching action via Odoo ORM (split: odoo part)
  3. Marks estado_actual to one of:
       APERTURA, NETEADO, COMISION_BANCARIA, HIPOTECA, GASTO_NO_DEDUCIBLE,
       COBRO_FACTURA, RECONCILED_OPEN_LINE, PENDIENTE_HUMANO, ERROR
  4. Re-uploads xlsx

Run: python3 dudas_apply.py [--company-vat B...] [--dry-run]
"""
# === pipeline isolation guard (auto-injected) ===
import os as _os, sys as _sys
_HERE = _os.path.dirname(_os.path.abspath(__file__))
if _HERE not in _sys.path:
    _sys.path.insert(0, _HERE)
try:
    import companies as _comp_guard
    if getattr(_comp_guard, "PIPELINE_NAME", None) != 'austral':
        raise RuntimeError(
            f"PIPELINE_MISMATCH: script {__file__} expected pipeline='austral' "
            f"but loaded companies.PIPELINE_NAME={getattr(_comp_guard, 'PIPELINE_NAME', None)!r}"
        )
except ImportError:
    pass  # script sin dependencia de companies.py (e.g. drive_ops)
# === end isolation guard ===

import argparse
import io
import json
import logging
import re
import sys
import subprocess
from pathlib import Path

sys.path.insert(0, _HERE)
import drive_ops  # noqa: E402
import companies as comp  # noqa: E402

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

XLSX_NAME = "dudas_para_revisar.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
HEADER = [
    "empresa", "tipo", "id_odoo", "ref_o_concepto", "fecha", "importe",
    "descripcion_corta", "motivo_duda", "sugerencia_actual",
    "tu_decision", "notas", "estado_actual", "primer_visto", "ultimo_visto",
]
ODOO_PYTHON = "/opt/odoo17/venv/bin/python"
APPLY_HELPER = "/opt/automation_austral/dudas_apply_odoo.py"

log = logging.getLogger("dudas_apply")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def classify_decision(decision: str) -> dict:
    """Map a free-text decision to a structured action."""
    d = (decision or "").lower().strip()
    if not d:
        return {"action": "skip", "label": "vacio"}

    if any(k in d for k in ["apertura", "asiento de apertura", "asiento apertura"]):
        return {"action": "skip", "label": "APERTURA"}
    if any(k in d for k in ["ignorar", "saltar"]):
        return {"action": "skip", "label": "IGNORADO"}

    if any(k in d for k in ["neteado", "neteo", "se compensa", "se anula con", "compensado", "siguiente", "anterior"]):
        return {"action": "skip", "label": "NETEADO"}

    if any(k in d for k in ["comision banc", "comisión banc", "comision bancaria", "comisión bancaria"]):
        return {"action": "direct_entry", "account_code": "626000", "label": "COMISION_BANCARIA"}

    if any(k in d for k in ["hipoteca", "préstamo", "prestamo"]):
        return {"action": "direct_entry", "account_code": "520000", "label": "HIPOTECA"}

    if any(k in d for k in ["no deducible", "no-deducible", "non deducible", "no deduc"]):
        return {"action": "direct_entry", "account_code": "629000", "label": "GASTO_NO_DEDUCIBLE",
                "narration": "Gasto NO deducible: " + decision[:200]}

    if any(k in d for k in ["pago factura de ingres", "pago factura del cliente", "cobro factura"]):
        return {"action": "match_open_aml", "account_code": "430000", "label": "COBRO_FACTURA"}

    if "pago seguro" in d or "seguro anual" in d:
        return {"action": "direct_entry", "account_code": "625000", "label": "SEGURO"}

    if any(k in d for k in [
        "falta fact", "falta fra", "falta f.r.a", "pendiente fact", "queda como duda",
        "queda el moviemiento", "queda el movimiento", "mientras se sube", "espera factura",
    ]):
        return {"action": "skip", "label": "PENDIENTE_FACTURA"}

    if "pago nomina" in d or "pago nómina" in d:
        return {"action": "partial_reconcile_465", "label": "PARTIAL_RECONCILE"}

    if "saldo pendiente" in d or "queda pendiente" in d or "diferencia se queda" in d:
        return {"action": "partial_reconcile", "label": "PARTIAL_RECONCILE"}

    if d in ("ok", "si", "sí", "confirmo", "vale", "correcto") or d.startswith("ok "):
        return {"action": "confirm_proposal", "label": "PARTIAL_RECONCILE"}

    if any(k in d for k in ["pago seguridad social", "pago ss", "seguridad social", "cotizacion ss", "tgss cotizacion"]):
        return {"action": "direct_entry", "account_code": "476000", "label": "PAGO_SS",
                "narration": "Pago Seg Social: " + decision[:200]}

    if any(k in d for k in ["pago iva", "liquidacion iva", "liquidación iva", "liquidacion de iva", "liquidación de iva", "autoliquidacion iva", "iva autoliquidacion", "modelo 303"]):
        return {"action": "direct_entry", "account_code": "477000", "label": "PAGO_IVA",
                "narration": "Pago liquidacion IVA: " + decision[:200]}

    if any(k in d for k in ["retenciones irpf", "pago irpf", "retenciones e ing", "retenciones a cta", "mod 111", "mod 115", "retenciones e ingresos a cuenta"]):
        return {"action": "direct_entry", "account_code": "475100", "label": "PAGO_IRPF",
                "narration": "Pago retenciones IRPF: " + decision[:200]}

    if any(k in d for k in ["pago alquiler", "alquiler mensual", "renta alquiler", "arrendamiento"]):
        return {"action": "match_open_aml", "account_code": "410000", "label": "PAGO_ALQUILER",
                "narration": "Pago alquiler: " + decision[:200]}

    # Usuario gestionará manualmente — dejar sin tocar
    if any(k in d for k in ["no hacer", "buscar la contrapartida", "buscare la contrapartida", "buscaré la contrapartida", "buscar contrapartida", "manualmente", "yo lo hago"]):
        return {"action": "skip", "label": "PENDIENTE_USUARIO"}

    # Usuario ha subido la factura — esperar a la próxima pasada
    if any(k in d for k in ["subida fact", "subida fra", "subido fact", "subido fra", "subida f.", "sube fact", "sube fra", "cubida fact", "factura subida"]):
        return {"action": "smart_subida_match", "label": "FACTURA_SUBIDA",
                "narration": "Factura subida, intentando match: " + decision[:200]}

    # Pago contra proveedor (factura antigua, distinto ejercicio)
    if any(k in d for k in ["contabiliza el pago contra el proveedor", "pago contra proveedor", "pago contra el proveedor", "pago a proveedor"]):
        return {"action": "match_open_aml", "account_code": "410000", "label": "PAGO_PROVEEDOR",
                "narration": decision[:200]}

    if any(k in d for k in ["pago factura", "pago fra", "pago de fra", "pago de factura",
                            "agrupacion de facturas", "agrupación de facturas",
                            "agrupacion facturas", "factura rectificativa", "factgura rectificativa"]):
        return {"action": "match_open_aml", "account_code": "410000", "label": "PAGO_FACTURA",
                "narration": "Pago factura proveedor: " + decision[:200]}

    # Cobros TPV / liquidaciones — cliente paga via TPV
    if "liquidacion efectuada" in d or "liquidaci" in d and "efectuada" in d or "cobro tpv" in d:
        return {"action": "direct_entry", "account_code": "430000", "label": "COBRO_TPV",
                "narration": "Cobro TPV: " + decision[:200]}

    # Gympass
    if "gympass" in d:
        return {"action": "direct_entry", "account_code": "430000", "label": "COBRO_GYMPASS",
                "partner_name": "Gympass US LLC",
                "narration": "Cobro Gympass: " + decision[:200]}

    # Comisiones bancarias / SEPA / TPV / devoluciones bancarias
    if any(k in d for k in [
        "gastos devolucion", "gastos devoluciones",
        "comisiones banc", "comision banc",
        "comsiones banc",
        "tarifa plana",
        "emision sepa", "emisi" "sepa",
        "emision remesa sepa",
        "liquidacion por emision",
        "liquidaci" "n por emisi",
    ]):
        return {"action": "direct_entry", "account_code": "626000", "label": "COMISION_BANCARIA",
                "narration": decision[:200]}

    # Devolucion de recibos cliente
    if "devolucion de recibo" in d or "devoluci" in d and "recibo" in d:
        return {"action": "direct_entry", "account_code": "430000", "label": "DEVOLUCION_CLIENTE",
                "narration": "Devolucion recibo SEPA: " + decision[:200]}

    # Emision Remesa SEPA: cobro masivo de clientes
    if "emision remesa" in d or "emisi" in d and "remesa" in d or "abona la cuenta de clientes" in d:
        return {"action": "direct_entry", "account_code": "430000", "label": "COBRO_REMESA_SEPA",
                "narration": "Cobro remesa SEPA: " + decision[:200]}

    return {"action": "human", "label": "PENDIENTE_HUMANO"}


def collect_rows(svc, cfg) -> tuple[str, list[dict], list]:
    folder = cfg.get("pending_folder")
    q = f"'{folder}' in parents and trashed=false and name='{XLSX_NAME}'"
    files = svc.files().list(q=q, fields="files(id,name)", supportsAllDrives=True).execute().get("files", [])
    if not files:
        return None, [], []
    fid = files[0]["id"]
    raw = svc.files().get_media(fileId=fid, supportsAllDrives=True).execute()
    wb = load_workbook(io.BytesIO(raw))
    ws = wb.active
    all_rows = list(ws.iter_rows(min_row=2, values_only=True))
    rows_with_decision = []
    for idx, row in enumerate(all_rows):
        if row is None or row[0] is None:
            continue
        d = (row[9] or "").strip() if row[9] else ""
        if not d:
            continue
        rows_with_decision.append({
            "row_index": idx,
            "empresa": row[0], "tipo": row[1], "id_odoo": row[2],
            "ref_o_concepto": row[3], "fecha": row[4], "importe": row[5],
            "descripcion_corta": row[6], "motivo_duda": row[7],
            "sugerencia_actual": row[8], "tu_decision": d,
            "notas": row[10] or "", "estado_actual": row[11] or "",
            "primer_visto": row[12] or "", "ultimo_visto": row[13] or "",
        })
    return fid, rows_with_decision, all_rows


def write_xlsx(rows_data: list[list]) -> bytes:
    wb = Workbook(); ws = wb.active; ws.title = "Dudas"
    ws.append(HEADER)
    hdr_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill; cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for col, w in {"A":25,"B":14,"C":9,"D":25,"E":12,"F":12,"G":30,"H":35,"I":35,"J":25,"K":40,"L":18,"M":12,"N":12}.items():
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"

    applied_fill = PatternFill(start_color="c6efce", end_color="c6efce", fill_type="solid")
    pending_fill = PatternFill(start_color="ffeb9c", end_color="ffeb9c", fill_type="solid")
    error_fill = PatternFill(start_color="ffc7ce", end_color="ffc7ce", fill_type="solid")
    for r in rows_data:
        ws.append(r)
        last = ws.max_row
        estado = (r[11] if len(r) > 11 else "") or ""
        if estado.startswith("ERROR"):
            for c in ws[last]: c.fill = error_fill
        elif estado in ("APERTURA","IGNORADO","NETEADO","COMISION_BANCARIA","HIPOTECA","GASTO_NO_DEDUCIBLE","COBRO_FACTURA","RECONCILED_OPEN_LINE","PARTIAL_RECONCILE","SEGURO","PENDIENTE_FACTURA"):
            for c in ws[last]: c.fill = applied_fill
        elif estado == "PENDIENTE_HUMANO":
            for c in ws[last]: c.fill = pending_fill
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def upload_xlsx(svc, fid: str, content: bytes):
    from googleapiclient.http import MediaIoBaseUpload
    media = MediaIoBaseUpload(io.BytesIO(content), mimetype=XLSX_MIME, resumable=False)
    svc.files().update(fileId=fid, media_body=media, supportsAllDrives=True).execute()



def classify_rechazo_decision(decision: str) -> dict:
    """Map free-text decision on Rechazados sheet to action."""
    d = (decision or "").lower().strip()
    if not d:
        return {"action": "skip", "label": ""}

    if any(k in d for k in ["borrar", "eliminar", "tirar", "delete", "papelera"]):
        return {"action": "rechazo_borrar", "label": "BORRADO"}

    if any(k in d for k in ["ignorar", "omitir", "saltar", "dejar"]):
        return {"action": "rechazo_ignorar", "label": "IGNORADO"}

    if any(k in d for k in ["reprocesar", "volver a procesar", "intentar de nuevo", "intenta", "vuelve a procesar", "reintentar"]):
        return {"action": "rechazo_reprocesar", "label": "REPROCESAR"}

    if any(k in d for k in ["rechazar", "rechazada", "rechazado", "archivar", "a rechazadas", "carpeta rechazadas", "no contabiliza"]):
        return {"action": "rechazo_archivar", "label": "ARCHIVADO_RECHAZADAS"}

    # Detect CIF/NIF Spanish format: optional letter + 7-8 digits + optional letter
    import re
    m = re.search(r"\b([a-z]?\d{7,8}[a-z]?)\b", d)
    has_vat_keyword = any(k in d for k in ["cif", "nif", "vat", "es el"])
    if m and (has_vat_keyword or len(d) < 80):
        vat = m.group(1).upper()
        return {"action": "rechazo_cif", "label": "CIF_CORREGIDO", "vat": vat}

    return {"action": "human", "label": "PENDIENTE_HUMANO_RECHAZO"}


def collect_rechazos(svc, cfg) -> list[dict]:
    """Read the Rechazados sheet of the company xlsx. Returns list of decisions."""
    folder = cfg.get("pending_folder")
    q = f"'{folder}' in parents and trashed=false and name='{XLSX_NAME}'"
    files = svc.files().list(q=q, fields="files(id,name)", supportsAllDrives=True).execute().get("files", [])
    if not files:
        return []
    fid = files[0]["id"]
    raw = svc.files().get_media(fileId=fid, supportsAllDrives=True).execute()
    wb = load_workbook(io.BytesIO(raw))
    if "Rechazados" not in wb.sheetnames:
        return []
    ws = wb["Rechazados"]
    out = []
    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or not row[0]:
            continue
        archivo, file_id, motivo, tu_decision, notas = (row + (None, None, None, None, None))[:5]
        decision = (tu_decision or "").strip() if tu_decision else ""
        if not decision:
            continue
        out.append({
            "row_index": idx,
            "archivo": archivo,
            "drive_file_id": file_id,
            "motivo": motivo,
            "tu_decision": decision,
            "notas": notas or "",
        })
    return out



def _do_drive_action(svc, action: dict, log_obj=None) -> dict:
    """Execute Drive ops for rechazo_borrar / _ignorar / _reprocesar. Returns result dict."""
    import re as _re
    file_id = action.get("drive_file_id")
    res = {"row_index": action.get("row_index"), "archivo": action.get("archivo"),
           "decision": action.get("decision"), "label": action.get("label")}
    if not file_id:
        res.update({"estado_actual": "ERROR_RECHAZO", "note": "drive_file_id ausente"})
        return res
    try:
        meta = svc.files().get(fileId=file_id, fields="parents,name", supportsAllDrives=True).execute()
        if not meta.get("parents"):
            res.update({"estado_actual": "ERROR_RECHAZO", "note": "archivo sin parents"})
            return res
        parent = meta["parents"][0]
        rev_meta = svc.files().get(fileId=parent, fields="parents", supportsAllDrives=True).execute()
        root = rev_meta["parents"][0] if rev_meta.get("parents") else parent

        act = action["action"]
        if act == "rechazo_borrar":
            svc.files().update(fileId=file_id, body={"trashed": True}, supportsAllDrives=True).execute()
            res.update({"estado_actual": "BORRADO", "note": "movido a papelera Drive"})
            return res

        if act == "rechazo_ignorar":
            q = "'%s' in parents and name='ignorados' and mimeType='application/vnd.google-apps.folder' and trashed=false" % root
            sub = svc.files().list(q=q, fields="files(id)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute().get("files", [])
            if sub:
                ign_id = sub[0]["id"]
            else:
                ign_id = svc.files().create(
                    body={"name": "ignorados", "mimeType": "application/vnd.google-apps.folder", "parents": [root]},
                    fields="id", supportsAllDrives=True,
                ).execute()["id"]
            svc.files().update(fileId=file_id, addParents=ign_id, removeParents=parent, supportsAllDrives=True).execute()
            res.update({"estado_actual": "IGNORADO", "note": "movido a ignorados/"})
            return res

        if act in ("rechazo_reprocesar", "rechazo_cif"):
            svc.files().update(fileId=file_id, addParents=root, removeParents=parent, supportsAllDrives=True).execute()
            res.update({"estado_actual": "REPROCESAR", "note": "movido a Pendientes/ — extractor lo procesará en la próxima pasada"})
            return res

        if act == "rechazo_archivar":
            # Regla global: facturas rechazadas definitivamente -> carpeta 'rechazadas' (per-company)
            # Resolver company por root (= pending_folder de la company)
            from companies import COMPANIES
            cfg = next((c for c in COMPANIES if c.get("pending_folder") == root), None) or {}
            rech_id = cfg.get("rechazadas_folder")
            if not rech_id:
                # fallback: buscar/crear 'rechazadas' como hermana de revision/
                q = "'%s' in parents and name='rechazadas' and mimeType='application/vnd.google-apps.folder' and trashed=false" % root
                sub = svc.files().list(q=q, fields="files(id)", supportsAllDrives=True, includeItemsFromAllDrives=True).execute().get("files", [])
                if sub:
                    rech_id = sub[0]["id"]
                else:
                    rech_id = svc.files().create(
                        body={"name": "rechazadas", "mimeType": "application/vnd.google-apps.folder", "parents": [root]},
                        fields="id", supportsAllDrives=True,
                    ).execute()["id"]
            svc.files().update(fileId=file_id, addParents=rech_id, removeParents=parent, supportsAllDrives=True).execute()
            res.update({"estado_actual": "ARCHIVADO_RECHAZADAS", "note": "movido a rechazadas/ — no se contabiliza"})
            return res

        res.update({"estado_actual": "ERROR_RECHAZO", "note": f"acción desconocida {act}"})
        return res
    except Exception as e:
        res.update({"estado_actual": "ERROR_RECHAZO", "note": f"exception: {str(e)[:200]}"})
        return res


def _extract_partner_from_motivo(motivo: str) -> str:
    import re as _re
    m = _re.search(r"contacto \[([^\]]+)\]", motivo or "")
    return (m.group(1).strip() if m else "").upper()


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--company-vat")
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    svc = drive_ops._service()
    overall = []

    for cfg in comp.COMPANIES:
        if args.company_vat and cfg["vat"] != args.company_vat:
            continue
        log.info(f"=== {cfg['name']} ===")
        fid, decisions, all_rows = collect_rows(svc, cfg)
        if not fid:
            log.warning(f"no xlsx for {cfg['name']}")
            continue

        # Classify decisions
        actions_payload = []
        for row in decisions:
            cls = classify_decision(row["tu_decision"])
            actions_payload.append({
                "row_index": row["row_index"],
                "tipo": row["tipo"], "id_odoo": row["id_odoo"],
                "importe": row["importe"], "decision": row["tu_decision"],
                "action": cls["action"], "label": cls["label"],
                "account_code": cls.get("account_code"),
                "narration": cls.get("narration"),
                "partner_name": cls.get("partner_name"),
                "sugerencia_actual": row.get("sugerencia_actual") or "",
                "ref_o_concepto": row.get("ref_o_concepto") or "",
            })

        log.info(f"  {len(decisions)} decisions -> classified")
        for a in actions_payload[:20]:
            log.info(f"    row#{a['row_index']} tipo={a['tipo']} id={a['id_odoo']} -> {a['label']}")

        # ===== RECHAZOS sheet =====
        rechazos = collect_rechazos(svc, cfg)
        rechazo_results = []
        for r in rechazos:
            cls = classify_rechazo_decision(r["tu_decision"])
            action = {
                "row_index": r["row_index"],
                "archivo": r["archivo"],
                "drive_file_id": r["drive_file_id"],
                "motivo": r["motivo"],
                "decision": r["tu_decision"],
                "action": cls["action"],
                "label": cls["label"],
                "vat": cls.get("vat"),
            }
            # Drive op (works in this venv since we have google libs)
            if cls["action"].startswith("rechazo_"):
                drive_res = _do_drive_action(svc, action)
                rechazo_results.append(drive_res)
                # For rechazo_cif, also queue an ORM action to create learned.rule + update partner
                if cls["action"] == "rechazo_cif":
                    partner_name = _extract_partner_from_motivo(r["motivo"])
                    actions_payload.append({
                        "row_index": r["row_index"],
                        "action": "create_vat_correction",
                        "label": "VAT_CORRECTION",
                        "partner_name": partner_name,
                        "vat": cls.get("vat"),
                        "company_id": cfg["odoo_company_id"],
                        "archivo": r["archivo"],
                    })
        if rechazo_results:
            log.info(f"  {len(rechazo_results)} rechazo decisions processed (Drive ops)")
            for r in rechazo_results:
                log.info(f"    {r['archivo']} -> {r['estado_actual']}: {r['note']}")

        # Hand off to Odoo helper to execute
        if not args.dry_run:
            tmp_in = Path(f"/tmp/dudas_austral/{cfg['vat']}_actions.json")
            tmp_out = Path(f"/tmp/dudas_austral/{cfg['vat']}_actions_result.json")
            tmp_in.parent.mkdir(parents=True, exist_ok=True)
            tmp_in.write_text(json.dumps({
                "company_id": cfg["odoo_company_id"],
                "company_vat": cfg["vat"],
                "actions": actions_payload,
            }, ensure_ascii=False))
            try:
                result = subprocess.run(
                    [ODOO_PYTHON, APPLY_HELPER, "--input", str(tmp_in), "--output", str(tmp_out)],
                    capture_output=True, text=True, timeout=600,
                )
                log.info(f"  helper rc={result.returncode}")
                if result.stderr:
                    log.info(f"  stderr tail: {result.stderr[-400:]}")
            except subprocess.TimeoutExpired:
                log.error("helper timed out")
            try:
                exec_results = json.loads(tmp_out.read_text())
            except Exception:
                exec_results = []
        else:
            exec_results = []

        results_by_idx = {r["row_index"]: r for r in exec_results}

        # Update xlsx rows in-place
        new_rows_data = []
        for idx, original in enumerate(all_rows):
            if original is None or original[0] is None:
                continue
            row_list = list(original)
            while len(row_list) < len(HEADER):
                row_list.append("")
            executed = results_by_idx.get(idx)
            if executed:
                row_list[11] = executed.get("estado_actual") or row_list[11]
                if executed.get("note"):
                    existing = row_list[10] or ""
                    row_list[10] = (existing + " | " + executed["note"])[:500] if existing else executed["note"][:500]
            new_rows_data.append(row_list)

        # NO sobrescribir el xlsx aquí — la republicación final con todas las hojas
        # (Dudas + Duplicados + Rechazados + Gastos_periodicos) es responsabilidad de
        # dudas_xlsx_publish.py que corre a las 23:39. Si lo hiciéramos aquí se perderían
        # las hojas extra y las tu_decision que el usuario había escrito en Rechazados.
        if False and not args.dry_run:
            upload_xlsx(svc, fid, write_xlsx(new_rows_data))

        overall.append({
            "company": cfg["name"],
            "decisions": len(decisions),
            "executed": len(exec_results),
            "labels": {a["label"]: sum(1 for x in actions_payload if x["label"] == a["label"]) for a in actions_payload},
        })

    print(json.dumps(overall, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
