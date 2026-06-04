#!/usr/bin/env python3
"""
Append-only xlsx generator for the dudas-to-review file in each company's root.

Looks for `dudas_para_revisar.xlsx` in the company's pending folder.
- If it doesn't exist, logs a warning (user creates it once).
- If it exists, downloads, merges new dudas (preserves your `tu_decision`
  column from existing rows), uploads back.

Key columns:
  empresa, tipo, id_odoo, ref_o_concepto, fecha, importe,
  descripcion_corta, motivo_duda, sugerencia_actual, tu_decision, notas,
  estado_actual, primer_visto, ultimo_visto
"""
# === pipeline isolation guard (auto-injected) ===
import os as _os, sys as _sys
_HERE = _os.path.dirname(_os.path.abspath(__file__))
if _HERE not in _sys.path:
    _sys.path.insert(0, _HERE)
try:
    import companies as _comp_guard
    if getattr(_comp_guard, "PIPELINE_NAME", None) != 'austral':
        raise RuntimeError(
            f"PIPELINE_MISMATCH: script {__file__} expected pipeline='austral' "
            f"but loaded companies.PIPELINE_NAME={getattr(_comp_guard, 'PIPELINE_NAME', None)!r}"
        )
except ImportError:
    pass  # script sin dependencia de companies.py (e.g. drive_ops)
# === end isolation guard ===

import argparse
import io
import json
import logging
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

ODOO_PATH = "/opt/odoo17/odoo"
ODOO_CONF = "/etc/odoo17.conf"
DB_NAME = "cararjfam_test"

sys.path.insert(0, ODOO_PATH)
sys.path.insert(0, _HERE)
import odoo  # noqa: E402
from odoo.api import Environment  # noqa: E402

import companies as comp  # noqa: E402

log = logging.getLogger("dudas_xlsx")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

XLSX_NAME = "dudas_para_revisar.xlsx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

HEADER = [
    "empresa", "tipo", "id_odoo", "ref_o_concepto", "fecha",
    "importe", "descripcion_corta", "motivo_duda", "sugerencia_actual",
    "tu_decision", "notas", "estado_actual", "primer_visto", "ultimo_visto",
]


def _confidence_from_narration(narr: str):
    if not narr:
        return None
    import re
    m = re.search(r"confianza[:\s]+([0-9]+\.?[0-9]*)", narr.lower())
    if not m:
        return None
    try:
        return float(m.group(1))
    except ValueError:
        return None


def _open_invoice_dudas(env, cid: int) -> list[dict]:
    drafts = env["account.move"].search([
        ("company_id", "=", cid),
        ("move_type", "in", ["in_invoice", "in_refund"]),
        ("state", "=", "draft"),
    ])
    out = []
    for m in drafts:
        conf = _confidence_from_narration(m.narration or "")
        reasons = []
        if conf is not None and conf < 0.9:
            reasons.append(f"confianza {conf:.2f}<0.9")
        if m.amount_total < 0:
            reasons.append("importe negativo (rectificativa?)")
        if not m.invoice_line_ids:
            reasons.append("sin lineas")
        if not reasons:
            reasons.append("revision general")
        sugg_codes = ",".join(sorted(set(l.account_id.code for l in m.invoice_line_ids if l.account_id)))
        out.append({
            "tipo": "factura",
            "id_odoo": m.id,
            "ref_o_concepto": m.ref or "",
            "fecha": str(m.invoice_date) if m.invoice_date else "",
            "importe": round(m.amount_total, 2),
            "descripcion_corta": (m.partner_id.name or "")[:80],
            "motivo_duda": "; ".join(reasons),
            "sugerencia_actual": f"cuenta lineas: {sugg_codes}" if sugg_codes else "",
            "estado_actual": m.state,
            "notas": (m.narration or "").replace("\n", " ").replace("\r", " ")[:300],
        })
    return out


def _open_bank_dudas(env, cid: int) -> list[dict]:
    import bank_matcher
    out = []
    for r in bank_matcher.propose_for_company(env, cid):
        non_rule = [x for x in r["proposals"] if x.get("kind") != "rule"]
        top = non_rule[0] if non_rule else None
        if top and top.get("score", 0) >= 90:
            continue
        sugg = "sin candidatos" if not top else f"{top.get('score',0)}% {top.get('rule_name') or top.get('move_name','?')} {top.get('partner','')}"
        reason = "multiples candidatos" if len(non_rule) > 1 else ("sin propuesta" if not top else "candidato unico bajo umbral")
        out.append({
            "tipo": "banco",
            "id_odoo": r["line_id"],
            "ref_o_concepto": (r["concept"] or "")[:100],
            "fecha": r["date"],
            "importe": r["amount"],
            "descripcion_corta": r["journal"],
            "motivo_duda": reason,
            "sugerencia_actual": sugg,
            "estado_actual": "no_conciliado",
            "notas": "",
        })
    near = bank_matcher.find_near_matches_for_company(env, cid)
    for r in near:
        n = r["near"]
        out.append({
            "tipo": "banco_descuadre",
            "id_odoo": r["bank_line_id"],
            "ref_o_concepto": (r["bank_concept"] or "")[:100],
            "fecha": r["bank_date"],
            "importe": r["bank_amount"],
            "descripcion_corta": f"{n['aml_account']} {n['aml_partner']}"[:80],
            "motivo_duda": f"diferencia {n['diff']:+.2f}€ con apunte abierto",
            "sugerencia_actual": f"apunte {n['aml_move_name']} = {n['aml_amount']:.2f}",
            "estado_actual": "descuadrado",
            "notas": (n['aml_label'] or "")[:300],
        })
    return out


def collect_for_company(env, cfg: dict) -> list[dict]:
    cid = cfg["odoo_company_id"]
    rows = []
    rows.extend(_open_invoice_dudas(env, cid))
    rows.extend(_open_bank_dudas(env, cid))
    for r in rows:
        r["empresa"] = cfg["name"]
    return rows


def _row_key(r) -> str:
    return f"{r.get('tipo','')}::{r.get('id_odoo','')}::{r.get('ref_o_concepto','')[:40]}"


def _row_to_list(r, today: str, primer_visto: str = ""):
    return [
        r.get("empresa", ""),
        r.get("tipo", ""),
        r.get("id_odoo", ""),
        r.get("ref_o_concepto", ""),
        r.get("fecha", ""),
        r.get("importe", ""),
        r.get("descripcion_corta", ""),
        r.get("motivo_duda", ""),
        r.get("sugerencia_actual", ""),
        r.get("tu_decision", ""),
        r.get("notas", ""),
        r.get("estado_actual", ""),
        primer_visto or today,
        today,
    ]


def _list_to_dict(row_values) -> dict:
    return {HEADER[i]: row_values[i] if i < len(row_values) else "" for i in range(len(HEADER))}


def find_xlsx(svc, folder_id: str):
    q = (f"'{folder_id}' in parents and trashed=false "
         f"and name='{XLSX_NAME}'")
    files = svc.files().list(q=q, fields="files(id,name)", supportsAllDrives=True).execute().get("files", [])
    return files[0] if files else None


def _build_workbook(rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dudas"
    ws.append(HEADER)
    hdr_fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center")
    for w_col, w in zip("ABCDEFGHIJKLMN", [25, 14, 9, 25, 12, 12, 30, 35, 35, 25, 40, 16, 12, 12]):
        ws.column_dimensions[w_col].width = w
    ws.freeze_panes = "A2"

    review_fill = PatternFill(start_color="fff2cc", end_color="fff2cc", fill_type="solid")
    descuadre_fill = PatternFill(start_color="fce4d6", end_color="fce4d6", fill_type="solid")
    for r in rows:
        ws.append(r)
        last = ws.max_row
        tipo_val = r[1] if len(r) > 1 else ""
        if tipo_val == "banco_descuadre":
            for cell in ws[last]:
                cell.fill = descuadre_fill
        elif r[9]:  # tu_decision filled
            for cell in ws[last]:
                cell.fill = review_fill

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def process_company(svc, env, cfg: dict, today: str) -> dict:
    folder = cfg.get("pending_folder")
    if not folder:
        return {"company": cfg["name"], "skipped": "no pending_folder"}

    file_meta = find_xlsx(svc, folder)
    if not file_meta:
        return {
            "company": cfg["name"],
            "skipped": (
                f"file '{XLSX_NAME}' not found in root of {cfg['name']} pending folder. "
                f"Crealo una vez (vacio) y el bot empezara a poblarlo."
            ),
        }

    existing_by_key = {}
    try:
        raw = svc.files().get_media(fileId=file_meta["id"], supportsAllDrives=True).execute()
        wb_existing = load_workbook(io.BytesIO(raw))
        ws_existing = wb_existing.active
        for row_cells in ws_existing.iter_rows(min_row=2, values_only=True):
            row_dict = _list_to_dict(list(row_cells))
            existing_by_key[_row_key(row_dict)] = row_dict
    except Exception as e:
        log.warning(f"[{cfg['name']}] could not read existing xlsx: {e}; will overwrite with fresh data")

    fresh_rows = collect_for_company(env, cfg)

    final_rows = []
    seen = set()
    new_count = 0
    for fr in fresh_rows:
        key = _row_key(fr)
        seen.add(key)
        existing = existing_by_key.get(key)
        if existing:
            fr_merged = dict(fr)
            fr_merged["tu_decision"] = existing.get("tu_decision", "") or ""
            primer = existing.get("primer_visto") or ""
            if not primer:
                primer = today
            user_notas = existing.get("notas") or ""
            if user_notas and len(user_notas) > len(fr.get("notas", "")):
                fr_merged["notas"] = user_notas
            final_rows.append(_row_to_list(fr_merged, today, primer))
        else:
            final_rows.append(_row_to_list(fr, today))
            new_count += 1

    closed_count = 0
    for k, ex in existing_by_key.items():
        if k in seen:
            continue
        if (ex.get("tu_decision") or "").strip():
            ex["estado_actual"] = "RESUELTO"
            final_rows.append(_row_to_list(ex, today, ex.get("primer_visto", "")))

    final_rows.sort(key=lambda r: (r[1] or "", str(r[2] or "")))
    xlsx_bytes = _build_workbook(final_rows)

    from googleapiclient.http import MediaIoBaseUpload
    media = MediaIoBaseUpload(io.BytesIO(xlsx_bytes), mimetype=XLSX_MIME, resumable=False)
    svc.files().update(fileId=file_meta["id"], media_body=media, supportsAllDrives=True).execute()

    return {
        "company": cfg["name"],
        "file_id": file_meta["id"],
        "rows_written": len(final_rows),
        "new_dudas": new_count,
        "preserved_decisions": sum(
            1 for r in final_rows if len(r) > 9 and (r[9] or "").strip()
        ),
        "closed_count": closed_count,
    }


def main():
    p = argparse.ArgumentParser()
    args = p.parse_args()

    odoo.tools.config.parse_config(["-c", ODOO_CONF])
    registry = odoo.registry(DB_NAME)
    import drive_ops
    svc = drive_ops._service()

    today = date.today().isoformat()
    out = []
    with registry.cursor() as cr:
        env = Environment(cr, odoo.SUPERUSER_ID, {"tz": "Europe/Madrid"})
        for cfg in comp.COMPANIES:
            try:
                stats = process_company(svc, env, cfg, today)
            except Exception as e:
                log.exception(f"[{cfg['name']}] failed")
                stats = {"company": cfg["name"], "error": str(e)}
            out.append(stats)
            log.info(f"[{cfg['name']}] {stats}")
    print(json.dumps(out, ensure_ascii=False, default=str))


if __name__ == "__main__":
    main()
